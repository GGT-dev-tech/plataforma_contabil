"""
=======================================================================
CONSOLIDADOR ANUAL  v1.0
=======================================================================
Autor   : Gerado via Antigravity (Google DeepMind)
Versão  : 1.0.0

COMO USAR:
  python3 consolidador_anual.py --ano 2026

  Lê o historico.json gerado pelo extrator_dre.py e produz:
    saida/2026/DRE_Anual_2026.xlsx

  Abas geradas:
    1. DRE Comparativo  — grupos como linhas, meses como colunas
    2. Evolução do Resultado — tabela mensal simplificada
    3. Em Aberto Acumulado  — pendências não quitadas no ano

Dependência: pip install openpyxl
=======================================================================
"""

import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Dependencia faltando. Execute:  pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────────

HIST_PATH = Path("historico.json")
DIR_SAIDA = Path("saida")

MESES_PT_ABREV = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
    5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}

MESES_PT = {
    1: "Janeiro",  2: "Fevereiro", 3: "Marco",    4: "Abril",
    5: "Maio",     6: "Junho",     7: "Julho",     8: "Agosto",
    9: "Setembro", 10: "Outubro",  11: "Novembro", 12: "Dezembro",
}

# Ordem e rótulos do DRE (mesmo do extrator)
LINHAS_DRE = [
    ("RECEITA_BRUTA",            "Receita Bruta",                      False, False),
    ("DEDUCOES_RECEITA",         "Deduções da Receita (Tributos)",      True,  False),
    ("__RL__",                   "( = ) RECEITA LIQUIDA",              False, True),
    ("CSP",                      "Custo dos Serviços Prestados",        True,  False),
    ("__LB__",                   "( = ) LUCRO BRUTO",                  False, True),
    ("DESPESAS_PESSOAL",         "Despesas com Pessoal",                True,  False),
    ("DESPESAS_ADMINISTRATIVAS", "Despesas Administrativas",            True,  False),
    ("DESPESAS_INSTALACOES",     "Despesas com Instalações",            True,  False),
    ("DESPESAS_COMERCIAIS",      "Despesas Comerciais",                 True,  False),
    ("__RO__",                   "( = ) RESULTADO OPERACIONAL",        False, True),
    ("OUTRAS_RECEITAS",          "Outras Receitas",                     False, False),
    ("__RAD__",                  "( = ) RESULTADO ANTES DISTRIBUICAO", False, True),
    ("DISTRIBUICAO_LUCROS",      "Distribuição de Lucros / Pró-labore", True,  False),
    ("__RF__",                   "( = ) RESULTADO FINAL",              False, True),
]

# ─────────────────────────────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────────────────────────────

COR_TITULO      = "1B2A4A"
COR_HEADER_MES  = "2C5F8A"
COR_LINHA_GRP   = "EBF5FB"
COR_SUBTOTAL    = "1B4F72"
COR_RESULTADO   = "0A1628"
COR_PAR         = "F5F8FA"
COR_IMPAR       = "FFFFFF"
COR_POS         = "1A6B34"
COR_NEG         = "8B1A1A"
COR_TOT_ANUAL   = "0D4A24"
COR_TOT_NEG     = "6B0E0E"
FMT_MOEDA       = '#,##0.00'

BORDA_FINA = Border(
    left=Side(style="thin", color="B0BEC5"),
    right=Side(style="thin", color="B0BEC5"),
    top=Side(style="thin", color="B0BEC5"),
    bottom=Side(style="thin", color="B0BEC5"),
)
BORDA_MED = Border(
    left=Side(style="medium", color="2C5F8A"),
    right=Side(style="medium", color="2C5F8A"),
    top=Side(style="medium", color="2C5F8A"),
    bottom=Side(style="medium", color="2C5F8A"),
)


def apf(c, cor):
    c.fill = PatternFill("solid", fgColor=cor)


def apft(c, bold=False, size=10, cor="000000", italic=False):
    c.font = Font(bold=bold, size=size, color=cor, italic=italic, name="Calibri")


def aln(c, h="left", v="center", wrap=False):
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def vc(ws, row, col, valor, bold=False, fundo=None, cor_f=None, med=False):
    """Célula de valor monetário."""
    c = ws.cell(row=row, column=col, value=valor if valor else None)
    c.number_format = FMT_MOEDA
    if fundo:
        apf(c, fundo)
    if cor_f is None:
        cor_f = COR_POS if (valor or 0) >= 0 else COR_NEG
    apft(c, bold=bold, size=10, cor=cor_f)
    aln(c, h="right")
    c.border = BORDA_MED if med else BORDA_FINA
    return c


# ─────────────────────────────────────────────────────────────────────
# LEITURA DO HISTÓRICO
# ─────────────────────────────────────────────────────────────────────

def ler_historico():
    if not HIST_PATH.exists():
        print(f"ERRO: {HIST_PATH} nao encontrado.")
        print("Execute primeiro:  python3 extrator_dre.py")
        sys.exit(1)
    with open(HIST_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def filtrar_por_ano(hist: dict, ano: int) -> dict:
    """Retorna apenas os meses do ano solicitado, ordenados."""
    meses = {}
    for chave, dados in hist.get("meses", {}).items():
        if dados.get("ano") == ano:
            mes = dados.get("mes", 0)
            meses[mes] = dados
    return dict(sorted(meses.items()))


def calcular_subtotais_mes(dados: dict) -> dict:
    """Retorna os subtotais calculados do historico.json de um mês."""
    return dados.get("subtotais", {
        "receita_liquida": 0,
        "lucro_bruto": 0,
        "resultado_operacional": 0,
        "resultado_antes_distribuicao": 0,
        "resultado_final": 0,
    })


def obter_valor_linha(chave: str, dados: dict) -> float:
    """Obtém o valor de uma linha do DRE para um mês específico."""
    if chave.startswith("__"):
        sub = calcular_subtotais_mes(dados)
        mapa_sub = {
            "__RL__" : sub.get("receita_liquida", 0),
            "__LB__" : sub.get("lucro_bruto", 0),
            "__RO__" : sub.get("resultado_operacional", 0),
            "__RAD__": sub.get("resultado_antes_distribuicao", 0),
            "__RF__" : sub.get("resultado_final", 0),
        }
        return mapa_sub.get(chave, 0)
    return dados.get("totais", {}).get(chave, 0)


# ─────────────────────────────────────────────────────────────────────
# ABA 1 — DRE COMPARATIVO
# ─────────────────────────────────────────────────────────────────────

def gerar_dre_comparativo(wb_out, meses_data: dict, ano: int):
    ws = wb_out.create_sheet("DRE Comparativo")
    ws.sheet_view.showGridLines = False

    meses_lista = sorted(meses_data.keys())  # [1, 2, 3, ...]
    n_meses     = len(meses_lista)
    col_desc    = 1
    col_meses   = {m: i + 2 for i, m in enumerate(meses_lista)}
    col_total   = n_meses + 2

    # ── Título ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=col_total)
    c = ws.cell(row=1, column=1,
                value=f"DRE COMPARATIVO — {ano}"
                      f"  |  {MESES_PT[meses_lista[0]]} a "
                      f"{MESES_PT[meses_lista[-1]]}")
    apf(c, COR_TITULO); apft(c, bold=True, size=13, cor="FFFFFF")
    aln(c, h="center"); ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=col_total)
    c = ws.cell(row=2, column=1,
                value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} "
                      f"| Regime de Caixa | Valores em R$")
    apf(c, "2C5F8A"); apft(c, italic=True, size=9, cor="E8F4FD")
    aln(c, h="center"); ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # ── Cabeçalhos de meses ─────────────────────────────────────────
    c = ws.cell(row=4, column=col_desc, value="DESCRIÇÃO")
    apf(c, COR_TITULO); apft(c, bold=True, size=10, cor="FFFFFF")
    aln(c, h="left"); c.border = BORDA_FINA

    for mes in meses_lista:
        col = col_meses[mes]
        c   = ws.cell(row=4, column=col,
                      value=MESES_PT_ABREV[mes])
        apf(c, COR_HEADER_MES); apft(c, bold=True, size=10, cor="FFFFFF")
        aln(c, h="center"); c.border = BORDA_FINA

    c = ws.cell(row=4, column=col_total, value="TOTAL ANO")
    apf(c, COR_TITULO); apft(c, bold=True, size=10, cor="FFFFFF")
    aln(c, h="center"); c.border = BORDA_FINA
    ws.row_dimensions[4].height = 22

    # ── Linhas do DRE ───────────────────────────────────────────────
    linha = 5
    for chave, rotulo, eh_despesa, eh_subtotal in LINHAS_DRE:
        fundo_linha = COR_PAR if linha % 2 == 0 else COR_IMPAR

        if eh_subtotal:
            fundo_linha = COR_SUBTOTAL if not chave.startswith("__RF__") \
                else COR_RESULTADO

        # Descrição
        c = ws.cell(row=linha, column=col_desc,
                    value=f"  {rotulo}" if not eh_subtotal else rotulo)
        apf(c, fundo_linha)
        apft(c, bold=eh_subtotal, size=10 if not eh_subtotal else 11,
             cor="FFFFFF" if eh_subtotal else "1B2A4A")
        aln(c, h="left"); c.border = BORDA_FINA if not eh_subtotal else BORDA_MED

        # Valores por mês
        total_ano = 0.0
        for mes in meses_lista:
            col   = col_meses[mes]
            dados = meses_data[mes]
            valor = obter_valor_linha(chave, dados)
            total_ano += valor
            vc(ws, linha, col, valor,
               bold=eh_subtotal,
               fundo=fundo_linha if not eh_subtotal else fundo_linha,
               cor_f="FFFFFF" if eh_subtotal else None,
               med=eh_subtotal)

        # Total do ano
        cor_tot = ("FFFFFF" if eh_subtotal
                   else COR_POS if total_ano >= 0 else COR_NEG)
        vc(ws, linha, col_total, total_ano,
           bold=True,
           fundo=fundo_linha,
           cor_f=cor_tot,
           med=True)

        ws.row_dimensions[linha].height = 20 if eh_subtotal else 17
        linha += 1

    # ── Larguras ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 44
    for mes in meses_lista:
        letter = get_column_letter(col_meses[mes])
        ws.column_dimensions[letter].width = 16
    ws.column_dimensions[get_column_letter(col_total)].width = 17

    # Congela
    ws.freeze_panes = "B5"


# ─────────────────────────────────────────────────────────────────────
# ABA 2 — EVOLUÇÃO DO RESULTADO
# ─────────────────────────────────────────────────────────────────────

def gerar_evolucao_resultado(wb_out, meses_data: dict, ano: int):
    ws = wb_out.create_sheet("Evolucao do Resultado")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1,
                value=f"EVOLUÇÃO DO RESULTADO — {ano}")
    apf(c, COR_TITULO); apft(c, bold=True, size=13, cor="FFFFFF")
    aln(c, h="center"); ws.row_dimensions[1].height = 28

    colunas = [
        "MES", "RECEITA BRUTA", "DEDUCOES", "RECEITA LIQUIDA",
        "CSP", "LUCRO BRUTO", "TOTAL DESPESAS", "RESULTADO OPERAC.",
        "DISTRIB. LUCROS", "RESULTADO FINAL", "SALDO BANCARIO",
        "ACUMULADO ANO",
    ]
    for i, col in enumerate(colunas, start=1):
        c = ws.cell(row=2, column=i, value=col)
        apf(c, COR_TITULO if i == 1 else COR_HEADER_MES)
        apft(c, bold=True, size=9, cor="FFFFFF")
        aln(c, h="center", wrap=True); c.border = BORDA_FINA
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 4

    meses_lista    = sorted(meses_data.keys())
    acum_corrente  = 0.0

    # Totais para linha de totais no final
    tot_rb = tot_ded = tot_rl = tot_csp = tot_lb = 0.0
    tot_td = tot_ro = tot_dl = tot_rf   = 0.0

    for i, mes in enumerate(meses_lista, start=4):
        dados   = meses_data[mes]
        totais  = dados.get("totais", {})
        sub     = calcular_subtotais_mes(dados)
        fundo   = COR_PAR if i % 2 == 0 else COR_IMPAR

        rb  = totais.get("RECEITA_BRUTA", 0)
        ded = totais.get("DEDUCOES_RECEITA", 0)
        rl  = sub.get("receita_liquida", 0)
        csp = totais.get("CSP", 0)
        lb  = sub.get("lucro_bruto", 0)
        dp  = totais.get("DESPESAS_PESSOAL", 0)
        da  = totais.get("DESPESAS_ADMINISTRATIVAS", 0)
        di  = totais.get("DESPESAS_INSTALACOES", 0)
        dc  = totais.get("DESPESAS_COMERCIAIS", 0)
        td  = dp + da + di + dc       # total despesas operacionais
        ro  = sub.get("resultado_operacional", 0)
        dl  = totais.get("DISTRIBUICAO_LUCROS", 0)
        rf  = sub.get("resultado_final", 0)
        sf  = dados.get("saldo_final", 0)

        acum_corrente += rf

        tot_rb  += rb;  tot_ded += ded; tot_rl += rl
        tot_csp += csp; tot_lb  += lb;  tot_td += td
        tot_ro  += ro;  tot_dl  += dl;  tot_rf += rf

        def fc(col, val, h="left"):
            c = ws.cell(row=i, column=col, value=val)
            apf(c, fundo); apft(c, size=9); aln(c, h=h); c.border = BORDA_FINA
            return c

        fc(1, MESES_PT_ABREV[mes], h="center")
        vc(ws, i, 2, rb,  fundo=fundo)
        vc(ws, i, 3, ded, fundo=fundo)
        vc(ws, i, 4, rl,  fundo=fundo)
        vc(ws, i, 5, csp, fundo=fundo)
        vc(ws, i, 6, lb,  fundo=fundo)
        vc(ws, i, 7, td,  fundo=fundo)
        vc(ws, i, 8, ro,  fundo=fundo)
        vc(ws, i, 9, dl,  fundo=fundo)
        vc(ws, i, 10, rf, fundo=fundo, med=True,
           cor_f="FFFFFF" if False else (COR_POS if rf >= 0 else COR_NEG))
        vc(ws, i, 11, sf, fundo=fundo)
        vc(ws, i, 12, acum_corrente, fundo=fundo,
           cor_f=COR_POS if acum_corrente >= 0 else COR_NEG)
        ws.row_dimensions[i].height = 18

    # ── Linha de totais ─────────────────────────────────────────────
    linha_tot = len(meses_lista) + 4
    ws.row_dimensions[linha_tot].height = 6  # separador
    linha_tot += 1

    ct = ws.cell(row=linha_tot, column=1, value="TOTAL ANUAL")
    apf(ct, COR_TITULO); apft(ct, bold=True, size=11, cor="FFFFFF")
    aln(ct, h="center"); ct.border = BORDA_MED

    for col, val in zip(range(2, 11),
                        [tot_rb, tot_ded, tot_rl, tot_csp, tot_lb,
                         tot_td, tot_ro, tot_dl, tot_rf]):
        vc(ws, linha_tot, col, val, bold=True,
           fundo=COR_TITULO, cor_f="FFFFFF", med=True)

    # Saldo do último mês (não soma)
    ult_mes = max(meses_lista)
    vc(ws, linha_tot, 11, meses_data[ult_mes].get("saldo_final", 0),
       bold=True, fundo=COR_TITULO, cor_f="FFFFFF", med=True)

    vc(ws, linha_tot, 12, tot_rf, bold=True,
       fundo=COR_RESULTADO, cor_f="FFFFFF", med=True)
    ws.row_dimensions[linha_tot].height = 26

    # Larguras
    ws.column_dimensions["A"].width = 8
    for i in range(2, 13):
        ws.column_dimensions[get_column_letter(i)].width = 16


# ─────────────────────────────────────────────────────────────────────
# ABA 3 — EM ABERTO ACUMULADO
# ─────────────────────────────────────────────────────────────────────

def gerar_em_aberto_acumulado(wb_out, meses_data: dict, ano: int):
    ws = wb_out.create_sheet("Em Aberto Acumulado")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
                value=f"LANÇAMENTOS EM ABERTO ACUMULADOS — {ano}")
    apf(c, COR_TITULO); apft(c, bold=True, size=13, cor="FFFFFF")
    aln(c, h="center"); ws.row_dimensions[1].height = 28

    colunas = ["MES DE ORIGEM", "DATA", "CATEGORIA", "VALOR (R$)", "RESUMO"]
    for i, col in enumerate(colunas, start=1):
        c = ws.cell(row=2, column=i, value=col)
        apf(c, COR_TITULO); apft(c, bold=True, size=10, cor="FFFFFF")
        aln(c, h="center"); c.border = BORDA_FINA
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 4

    linha       = 4
    total_aberto = 0.0

    for mes in sorted(meses_data.keys()):
        dados    = meses_data[mes]
        abertos  = dados.get("em_aberto_detalhes", [])
        if not abertos:
            continue

        for item in abertos:
            fundo = COR_PAR if linha % 2 == 0 else COR_IMPAR

            def fc(col, val, h="left", wrap=False):
                c = ws.cell(row=linha, column=col, value=val)
                apf(c, fundo); apft(c, size=9); aln(c, h=h, wrap=wrap)
                c.border = BORDA_FINA
                return c

            fc(1, MESES_PT.get(mes, str(mes)), h="center")
            fc(2, item.get("data", ""), h="center")
            fc(3, item.get("categoria", ""))
            vc(ws, linha, 4, item.get("valor", 0), fundo=fundo)
            fc(5, str(item.get("resumo", ""))[:80], wrap=True)
            total_aberto += item.get("valor", 0)
            ws.row_dimensions[linha].height = 15
            linha += 1

    if linha == 4:
        ws.merge_cells("A4:E4")
        c = ws.cell(row=4, column=1,
                    value="Nenhum lancamento em aberto registrado no período.")
        apft(c, italic=True, cor="888888"); aln(c, h="center")
    else:
        linha += 1
        ct = ws.cell(row=linha, column=1, value="TOTAL EM ABERTO")
        apf(ct, COR_TITULO); apft(ct, bold=True, size=11, cor="FFFFFF")
        aln(ct, h="left"); ct.border = BORDA_MED
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha, end_column=3)
        vc(ws, linha, 4, total_aberto, bold=True,
           fundo=COR_TITULO, cor_f="FFFFFF", med=True)
        ws.row_dimensions[linha].height = 24

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 50


# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Consolidador Anual DRE — lê historico.json e gera "
                    "planilha comparativa anual"
    )
    parser.add_argument(
        "--ano", type=int, default=datetime.now().year,
        help="Ano a consolidar (default: ano atual)"
    )
    args = parser.parse_args()
    ano  = args.ano

    print("=" * 60)
    print(f"  CONSOLIDADOR ANUAL  v1.0  —  Ano {ano}")
    print("=" * 60)

    hist       = ler_historico()
    meses_data = filtrar_por_ano(hist, ano)

    if not meses_data:
        print(f"\nNenhum dado encontrado para {ano} no historico.json.")
        print("Execute primeiro:  python3 extrator_dre.py")
        sys.exit(0)

    meses_lista = sorted(meses_data.keys())
    print(f"\n  {len(meses_lista)} mes(es) encontrado(s): "
          f"{', '.join(MESES_PT_ABREV[m] for m in meses_lista)}")

    dir_ano   = DIR_SAIDA / str(ano)
    dir_ano.mkdir(parents=True, exist_ok=True)
    arq_saida = dir_ano / f"DRE_Anual_{ano}.xlsx"

    wb_out = openpyxl.Workbook()
    if "Sheet" in wb_out.sheetnames:
        del wb_out["Sheet"]

    print("\n  -> Aba 1: DRE Comparativo ...")
    gerar_dre_comparativo(wb_out, meses_data, ano)

    print("  -> Aba 2: Evolução do Resultado ...")
    gerar_evolucao_resultado(wb_out, meses_data, ano)

    print("  -> Aba 3: Em Aberto Acumulado ...")
    gerar_em_aberto_acumulado(wb_out, meses_data, ano)

    wb_out.save(str(arq_saida))

    # Resumo rápido
    total_rf = sum(
        meses_data[m].get("subtotais", {}).get("resultado_final", 0)
        for m in meses_lista
    )
    print(f"\n  Resultado acumulado {ano}: R$ {total_rf:,.2f}")
    print(f"\n  Arquivo gerado: {arq_saida.resolve()}")
    print("=" * 60)


if __name__ == "__main__":
    main()
