import openpyxl
from decimal import Decimal
from datetime import datetime, date
from typing import List, Dict, Any
from app.models.domain import TipoStaging

class StandardTemplateParser:
    """
    Parser para a Planilha Padrão do Sistema com 4 abas:
    1_Receitas, 2_Despesas, 3_Extrato_Bancario, 4_Movimentacao_Dinheiro
    Alimenta a tabela staging_registros.
    """
    def parse(self, file_path_or_stream) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(file_path_or_stream, data_only=True)
        staging_items = []

        def parse_date(val):
            if isinstance(val, (datetime, date)):
                return val.strftime("%Y-%m-%d") if isinstance(val, datetime) else val.isoformat()
            if isinstance(val, str):
                val = val.strip()
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(val, fmt).date().isoformat()
                    except ValueError:
                        pass
            return datetime.utcnow().date().isoformat()

        def parse_decimal(val):
            if val is None: return Decimal("0.00")
            if isinstance(val, (int, float)): return Decimal(str(val))
            val_str = str(val).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
            try:
                return Decimal(val_str)
            except Exception:
                return Decimal("0.00")

        # 1. Receitas
        if "1_Receitas" in wb.sheetnames:
            ws = wb["1_Receitas"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                staging_items.append({
                    "tipo": TipoStaging.RECEITA,
                    "data": parse_date(row[0]),
                    "entidade_nome": str(row[1] or ""),
                    "descricao": str(row[2] or ""),
                    "valor": abs(parse_decimal(row[3])),
                    "forma_pagamento": str(row[4] or "") if len(row) > 4 else None,
                    "conta_destino": str(row[5] or "") if len(row) > 5 else None
                })

        # 2. Despesas
        if "2_Despesas" in wb.sheetnames:
            ws = wb["2_Despesas"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                staging_items.append({
                    "tipo": TipoStaging.DESPESA,
                    "data": parse_date(row[0]),
                    "entidade_nome": str(row[1] or ""),
                    "cnpj_cpf": str(row[2] or "") if len(row) > 2 else None,
                    "categoria": str(row[3] or "") if len(row) > 3 else None,
                    "descricao": str(row[3] or "") if len(row) > 3 else "Despesa Importada",
                    "valor": abs(parse_decimal(row[4])) if len(row) > 4 else Decimal("0.00"),
                    "conta_origem": str(row[5] or "") if len(row) > 5 else None
                })

        # 3. Extrato Bancario
        if "3_Extrato_Bancario" in wb.sheetnames:
            ws = wb["3_Extrato_Bancario"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                valor = parse_decimal(row[2]) if len(row) > 2 else Decimal("0.00")
                tipo_dc = str(row[3] or "D").upper() if len(row) > 3 else ("C" if valor > 0 else "D")
                if tipo_dc == "D" and valor > 0: valor = -valor
                staging_items.append({
                    "tipo": TipoStaging.EXTRATO,
                    "data": parse_date(row[0]),
                    "descricao": str(row[1] or ""),
                    "valor": valor,
                    "conta_origem": str(row[4] or "Banco") if len(row) > 4 else "Banco"
                })

        # 4. Movimentacao Dinheiro
        if "4_Movimentacao_Dinheiro" in wb.sheetnames:
            ws = wb["4_Movimentacao_Dinheiro"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                valor = parse_decimal(row[2]) if len(row) > 2 else Decimal("0.00")
                tipo_mov = str(row[3] or "Saida").lower() if len(row) > 3 else "saida"
                if "saida" in tipo_mov and valor > 0: valor = -valor
                staging_items.append({
                    "tipo": TipoStaging.DINHEIRO,
                    "data": parse_date(row[0]),
                    "descricao": str(row[1] or ""),
                    "valor": valor,
                    "conta_origem": "Caixa Geral (Dinheiro)",
                    "entidade_nome": str(row[4] or "") if len(row) > 4 else None
                })

        return staging_items
