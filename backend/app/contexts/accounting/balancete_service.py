import uuid
from typing import Dict, Any, List, Optional
from datetime import date
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.plano_contas import PlanoDeContas, GrupoConta, NaturezaConta
from app.models.ledger import LancamentoCabecalho, PartidaItem, TipoPartida, StatusLancamento

class BalanceteService:
    def __init__(self, db: Session):
        self.db = db

    def _to_uuid(self, val):
        if not val: return None
        if isinstance(val, uuid.UUID): return val
        try:
            return uuid.UUID(str(val))
        except Exception:
            return None

    def calcular_balancete(self, empresa_id: Any, dt_inicio: date, dt_fim: date) -> Dict[str, Any]:
        emp_uuid = self._to_uuid(empresa_id)

        # 1. Carregar Plano de Contas
        contas = self.db.query(PlanoDeContas).filter(
            PlanoDeContas.empresa_id == emp_uuid
        ).order_by(PlanoDeContas.codigo_contabil).all()

        # 2. Carregar todas as partidas contábeis do período
        partidas_periodo = self.db.query(PartidaItem).join(LancamentoCabecalho).filter(
            LancamentoCabecalho.empresa_id == emp_uuid,
            LancamentoCabecalho.status == StatusLancamento.CONFIRMADO,
            LancamentoCabecalho.data_competencia >= dt_inicio,
            LancamentoCabecalho.data_competencia <= dt_fim
        ).all()

        # 3. Agrupar Débitos e Créditos por Conta
        debitos_mes = {}
        creditos_mes = {}

        for p in partidas_periodo:
            conta_id = str(p.conta_contabil_id)
            val = float(p.valor or 0.0)
            if p.natureza == TipoPartida.DEBITO:
                debitos_mes[conta_id] = debitos_mes.get(conta_id, 0.0) + val
            else:
                creditos_mes[conta_id] = creditos_mes.get(conta_id, 0.0) + val

        # 4. Montar Linhas do Balancete
        linhas = []
        total_debitos = 0.0
        total_creditos = 0.0

        for c in contas:
            cid = str(c.id)
            deb_m = debitos_mes.get(cid, 0.0)
            cred_m = creditos_mes.get(cid, 0.0)

            saldo_anterior = float(getattr(c, 'saldo_inicial', 0.0) or 0.0)
            
            is_devedora = (c.natureza == NaturezaConta.DEVEDORA or str(c.grupo.value if hasattr(c.grupo, 'value') else c.grupo) in ["ATIVO", "DESPESA", "RESULTADO"])
            
            if is_devedora:
                saldo_atual = saldo_anterior + deb_m - cred_m
            else:
                saldo_atual = saldo_anterior + cred_m - deb_m

            if abs(saldo_anterior) < 0.01 and abs(deb_m) < 0.01 and abs(cred_m) < 0.01 and abs(saldo_atual) < 0.01:
                continue

            linha = {
                "codigo": c.codigo_contabil,
                "descricao": c.descricao,
                "grupo": c.grupo.value if hasattr(c.grupo, 'value') else str(c.grupo),
                "natureza": c.natureza.value if hasattr(c.natureza, 'value') else str(c.natureza),
                "saldo_anterior": saldo_anterior,
                "debito_mes": deb_m,
                "credito_mes": cred_m,
                "saldo_atual": saldo_atual,
            }
            linhas.append(linha)

            total_debitos += deb_m
            total_creditos += cred_m

        return {
            "periodo": f"{dt_inicio.strftime('%d/%m/%Y')} a {dt_fim.strftime('%d/%m/%Y')}",
            "linhas": linhas,
            "totais": {
                "total_debitos": total_debitos,
                "total_creditos": total_creditos,
                "diferenca_balanco": abs(total_debitos - total_creditos),
                "equilibrado": abs(total_debitos - total_creditos) < 0.01
            }
        }

    def exportar_excel(self, dados_balancete: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balancete Analítico"
        ws.sheet_view.showGridLines = True

        COR_NAVY = "1B2A4A"
        COR_HEADER = "2C5F8A"
        COR_PAR = "F0F4F8"

        font_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        font_sub = Font(name="Arial", size=10, italic=True, color="FFFFFF")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=10, bold=True)
        font_data = Font(name="Arial", size=10)

        fill_titulo = PatternFill("solid", fgColor=COR_NAVY)
        fill_header = PatternFill("solid", fgColor=COR_HEADER)
        fill_par = PatternFill("solid", fgColor=COR_PAR)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        border_thin = Border(
            left=Side(style="thin", color="B0BEC5"), right=Side(style="thin", color="B0BEC5"),
            top=Side(style="thin", color="B0BEC5"), bottom=Side(style="thin", color="B0BEC5")
        )

        ws.merge_cells("A1:G1")
        ws["A1"] = "BALANCETE ANALÍTICO CONTÁBIL"
        ws["A1"].font = font_titulo
        ws["A1"].fill = fill_titulo
        ws["A1"].alignment = align_center

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Período: {dados_balancete.get('periodo', '')}"
        ws["A2"].font = font_sub
        ws["A2"].fill = fill_titulo
        ws["A2"].alignment = align_center

        headers = ["Código Conta", "Descrição da Conta", "Grupo", "Saldo Anterior", "Débitos (Mês)", "Créditos (Mês)", "Saldo Atual"]
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin

        row_idx = 5
        for i, l in enumerate(dados_balancete.get("linhas", [])):
            fill_current = fill_par if i % 2 == 0 else PatternFill(fill_type=None)
            
            c1 = ws.cell(row=row_idx, column=1, value=l["codigo"])
            c2 = ws.cell(row=row_idx, column=2, value=l["descricao"])
            c3 = ws.cell(row=row_idx, column=3, value=l["grupo"])
            c4 = ws.cell(row=row_idx, column=4, value=l["saldo_anterior"])
            c5 = ws.cell(row=row_idx, column=5, value=l["debito_mes"])
            c6 = ws.cell(row=row_idx, column=6, value=l["credito_mes"])
            c7 = ws.cell(row=row_idx, column=7, value=l["saldo_atual"])

            c1.alignment = align_left
            c2.alignment = align_left
            c3.alignment = align_center
            c4.alignment = align_right
            c5.alignment = align_right
            c6.alignment = align_right
            c7.alignment = align_right

            c4.number_format = "#,##0.00;(#,##0.00);\"-\""
            c5.number_format = "#,##0.00;(#,##0.00);\"-\""
            c6.number_format = "#,##0.00;(#,##0.00);\"-\""
            c7.number_format = "#,##0.00;(#,##0.00);\"-\""

            for c in [c1, c2, c3, c4, c5, c6, c7]:
                c.font = font_data
                c.border = border_thin
                if fill_current.fill_type: c.fill = fill_current

            row_idx += 1

        totais = dados_balancete.get("totais", {})
        ws.cell(row=row_idx, column=1, value="TOTAL GERAL").font = font_bold
        ws.cell(row=row_idx, column=5, value=totais.get("total_debitos", 0.0)).font = font_bold
        ws.cell(row=row_idx, column=6, value=totais.get("total_creditos", 0.0)).font = font_bold

        ws.cell(row=row_idx, column=5).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=6).number_format = "#,##0.00"

        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = Border(top=Side(style="medium"), bottom=Side(style="double"))

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        import io
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
