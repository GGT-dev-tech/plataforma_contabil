from typing import Dict, Any, List
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime
from uuid import UUID
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models.ledger import LancamentoCabecalho, PartidaItem, TipoPartida, StatusLancamento
from app.models.plano_contas import PlanoDeContas, ClassificacaoDRE

class DREService:
    """
    Serviço responsável por gerar a Demonstração do Resultado do Exercício (DRE)
    baseada nas partidas contábeis efetivadas (regime de caixa/competência).
    """

    ESTRUTURA_DRE = [
        (ClassificacaoDRE.RECEITA_BRUTA, False, None),
        (ClassificacaoDRE.DEDUCOES_RECEITA, True, "RECEITA LIQUIDA"),
        (ClassificacaoDRE.CSP, True, "LUCRO BRUTO"),
        (ClassificacaoDRE.DESPESAS_PESSOAL, True, None),
        (ClassificacaoDRE.DESPESAS_ADMINISTRATIVAS, True, None),
        (ClassificacaoDRE.DESPESAS_INSTALACOES, True, None),
        (ClassificacaoDRE.DESPESAS_COMERCIAIS, True, "RESULTADO OPERACIONAL LIQUIDO"),
        (ClassificacaoDRE.OUTRAS_RECEITAS, False, "RESULTADO ANTES DA DISTRIBUICAO"),
        (ClassificacaoDRE.DISTRIBUICAO_LUCROS, True, "RESULTADO FINAL APOS DISTRIBUICAO"),
    ]

    ROTULOS_GRUPO = {
        ClassificacaoDRE.RECEITA_BRUTA: "Receita Bruta",
        ClassificacaoDRE.DEDUCOES_RECEITA: "Deduções da Receita (Tributos)",
        ClassificacaoDRE.CSP: "Custo dos Serviços Prestados",
        ClassificacaoDRE.DESPESAS_PESSOAL: "Despesas com Pessoal",
        ClassificacaoDRE.DESPESAS_ADMINISTRATIVAS: "Despesas Administrativas",
        ClassificacaoDRE.DESPESAS_INSTALACOES: "Despesas com Instalações",
        ClassificacaoDRE.DESPESAS_COMERCIAIS: "Despesas Comerciais",
        ClassificacaoDRE.OUTRAS_RECEITAS: "Outras Receitas",
        ClassificacaoDRE.DISTRIBUICAO_LUCROS: "Distribuição de Lucros / Pró-labore",
    }

    def __init__(self, db: Session):
        self.db = db

    def calcular_dre_periodo(self, empresa_id: UUID, dt_inicio: date, dt_fim: date) -> Dict[str, Any]:
        """
        Calcula o DRE de um período específico consolidando as partidas.
        No regime normal, Receitas (C) são positivas, Despesas (D) são negativas.
        Aqui extraímos o valor absoluto e estruturamos pelo ClassificacaoDRE.
        """
        # Consulta todas as partidas efetivadas no período que possuem classificação DRE
        resultados = (
            self.db.query(
                PlanoDeContas.classificacao_dre,
                func.sum(PartidaItem.valor).label("total")
            )
            .join(PartidaItem, PartidaItem.conta_contabil_id == PlanoDeContas.id)
            .join(LancamentoCabecalho, LancamentoCabecalho.id == PartidaItem.cabecalho_id)
            .filter(
                LancamentoCabecalho.empresa_id == empresa_id,
                LancamentoCabecalho.data_competencia >= dt_inicio,
                LancamentoCabecalho.data_competencia <= dt_fim,
                LancamentoCabecalho.status.in_([StatusLancamento.CONFIRMADO, StatusLancamento.EXPORTADO]),
                PlanoDeContas.classificacao_dre.isnot(None)
            )
            .group_by(PlanoDeContas.classificacao_dre)
            .all()
        )

        totais = {row.classificacao_dre: float(row.total or 0.0) for row in resultados}
        
        return self._estruturar_dre(totais)

    def _estruturar_dre(self, totais: Dict[ClassificacaoDRE, float]) -> Dict[str, Any]:
        """Organiza os totais brutos na estrutura em cascata do DRE."""
        linhas = []
        saldo = 0.0
        
        rb = totais.get(ClassificacaoDRE.RECEITA_BRUTA, 0.0)

        for grupo, subtrativo, label_sub in self.ESTRUTURA_DRE:
            vg = totais.get(grupo, 0.0)
            
            # Atualiza saldo
            if grupo == ClassificacaoDRE.RECEITA_BRUTA:
                saldo += vg
            elif grupo == ClassificacaoDRE.OUTRAS_RECEITAS:
                saldo += vg
            elif subtrativo:
                saldo -= abs(vg)
                
            pct = round((abs(vg) / rb * 100), 2) if rb else 0.0

            linhas.append({
                "grupo": grupo.name,
                "descricao": self.ROTULOS_GRUPO.get(grupo, grupo.name),
                "valor": vg,
                "percentual_receita": pct,
                "is_subtotal": False
            })

            if label_sub:
                pct_sub = round((abs(saldo) / rb * 100), 2) if rb else 0.0
                linhas.append({
                    "grupo": f"SUBTOTAL_{grupo.name}",
                    "descricao": f"( = ) {label_sub}",
                    "valor": round(saldo, 2),
                    "percentual_receita": pct_sub,
                    "is_subtotal": True
                })

        subtotais = {
            "receita_liquida": next((l["valor"] for l in linhas if l["descricao"] == "( = ) RECEITA LIQUIDA"), 0.0),
            "lucro_bruto": next((l["valor"] for l in linhas if l["descricao"] == "( = ) LUCRO BRUTO"), 0.0),
            "resultado_operacional": next((l["valor"] for l in linhas if l["descricao"] == "( = ) RESULTADO OPERACIONAL LIQUIDO"), 0.0),
            "resultado_antes_distribuicao": next((l["valor"] for l in linhas if l["descricao"] == "( = ) RESULTADO ANTES DA DISTRIBUICAO"), 0.0),
            "resultado_final": next((l["valor"] for l in linhas if l["descricao"] == "( = ) RESULTADO FINAL APOS DISTRIBUICAO"), 0.0),
        }

        return {
            "linhas": linhas,
            "subtotais": subtotais
        }

    def calcular_acumulado_ano(self, empresa_id: UUID, ano: int) -> Dict[str, Any]:
        """Calcula o consolidado anual (meses lado a lado + acumulado)."""
        dt_inicio_ano = date(ano, 1, 1)
        dt_fim_ano = date(ano, 12, 31)
        
        # Puxa o DRE consolidado do ano inteiro
        dre_anual = self.calcular_dre_periodo(empresa_id, dt_inicio_ano, dt_fim_ano)
        
        # Opcional: Gerar por mês para um relatório matricial
        meses = {}
        for mes in range(1, 13):
            # simplificação para fim do mês
            import calendar
            _, last_day = calendar.monthrange(ano, mes)
            dt_inicio_mes = date(ano, mes, 1)
            dt_fim_mes = date(ano, mes, last_day)
            meses[f"{ano}-{mes:02d}"] = self.calcular_dre_periodo(empresa_id, dt_inicio_mes, dt_fim_mes)
            
        return {
            "ano": ano,
            "acumulado": dre_anual,
            "meses": meses
        }

    def exportar_excel(self, dados_dre: Dict[str, Any], periodo: str) -> bytes:
        """
        Gera um arquivo Excel idêntico ao do extrator_dre original, 
        baseado no JSON estruturado do DRE.
        Retorna os bytes do arquivo para download via API.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DRE"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A8"

        # Estilos (reaproveitados do extrator)
        COR_TITULO_FUNDO   = "1B2A4A"
        COR_TITULO_FONTE   = "FFFFFF"
        COR_GRUPO_FUNDO    = "2C5F8A"
        COR_GRUPO_FONTE    = "FFFFFF"
        COR_LINHA_PAR      = "F0F4F8"
        COR_LINHA_IMPAR    = "FFFFFF"
        COR_POSITIVO       = "1A6B34"
        COR_NEGATIVO       = "8B1A1A"
        COR_RESULTADO_POS  = "0D4A24"
        COR_RESULTADO_NEG  = "6B0E0E"
        COR_HEADER_ABA     = "0A1628"

        BORDA_FINA = Border(
            left=Side(style="thin", color="B0BEC5"),
            right=Side(style="thin", color="B0BEC5"),
            top=Side(style="thin", color="B0BEC5"),
            bottom=Side(style="thin", color="B0BEC5"),
        )
        BORDA_MEDIA = Border(
            left=Side(style="medium", color="2C5F8A"),
            right=Side(style="medium", color="2C5F8A"),
            top=Side(style="medium", color="2C5F8A"),
            bottom=Side(style="medium", color="2C5F8A"),
        )

        def apf(cell, hex_cor):
            cell.fill = PatternFill("solid", fgColor=hex_cor)

        def apft(cell, bold=False, italic=False, size=10, cor="000000"):
            cell.font = Font(bold=bold, italic=italic, size=size, color=cor, name="Calibri")

        def aln(cell, h="left", v="center"):
            cell.alignment = Alignment(horizontal=h, vertical=v)

        def val_cell(row, col, valor, bold=False, fundo=None, cor_fonte=None, total=False):
            c = ws.cell(row=row, column=col, value=valor if valor != 0 else None)
            c.number_format = '#,##0.00'
            if fundo: apf(c, fundo)
            if not cor_fonte: cor_fonte = COR_POSITIVO if (valor or 0) >= 0 else COR_NEGATIVO
            apft(c, bold=bold or total, size=11 if total else 10, cor=cor_fonte)
            aln(c, h="right")
            c.border = BORDA_MEDIA if total else BORDA_FINA
            return c

        # Cabeçalho
        ws.merge_cells("A1:C1")
        c1 = ws.cell(row=1, column=1, value=f"DEMONSTRACAO DO RESULTADO DO EXERCICIO — {periodo}")
        apf(c1, COR_TITULO_FUNDO); apft(c1, bold=True, size=13, cor=COR_TITULO_FONTE); aln(c1, h="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:C2")
        c2 = ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Regime de Caixa/Competência via Plataforma Contábil")
        apf(c2, "2C5F8A"); apft(c2, italic=True, size=9, cor="E8F4FD"); aln(c2, h="center")
        ws.row_dimensions[2].height = 16

        # Headers de coluna
        for i, texto in enumerate(["DESCRICAO", "VALOR (R$)", "% RECEITA BRUTA"], start=1):
            c = ws.cell(row=6, column=i, value=texto)
            apf(c, COR_HEADER_ABA); apft(c, bold=True, cor=COR_TITULO_FONTE, size=10); aln(c, h="center")
            c.border = BORDA_FINA
        ws.row_dimensions[6].height = 22

        linha = 8
        linhas_dre = dados_dre.get("linhas", [])
        
        for item in linhas_dre:
            is_sub = item.get("is_subtotal", False)
            valor = item.get("valor", 0.0)
            pct = item.get("percentual_receita", 0.0)
            
            if is_sub:
                ws.row_dimensions[linha].height = 6
                linha += 1
                cf = COR_RESULTADO_POS if valor >= 0 else COR_RESULTADO_NEG
                cd = ws.cell(row=linha, column=1, value=item["descricao"])
                apf(cd, cf); apft(cd, bold=True, size=11, cor="FFFFFF"); aln(cd, h="left"); cd.border = BORDA_MEDIA
                
                val_cell(linha, 2, round(valor, 2), bold=True, fundo=cf, cor_fonte="FFFFFF", total=True)
                
                cpt = ws.cell(row=linha, column=3, value=f"{pct:.1f}%" if pct else "—")
                apf(cpt, cf); apft(cpt, bold=True, cor="FFFFFF", size=11); aln(cpt, h="center"); cpt.border = BORDA_MEDIA
                
                ws.row_dimensions[linha].height = 24
                linha += 2
            else:
                # Título do Grupo
                grp_c = ws.cell(row=linha, column=1, value=f"  {item['descricao']}")
                apf(grp_c, COR_GRUPO_FUNDO); apft(grp_c, bold=True, size=10, cor=COR_GRUPO_FONTE)
                aln(grp_c, h="left"); grp_c.border = BORDA_FINA
                ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
                ws.row_dimensions[linha].height = 20
                linha += 1
                
                # Linha de total
                fundo = COR_LINHA_PAR if linha % 2 == 0 else COR_LINHA_IMPAR
                c = ws.cell(row=linha, column=1, value=f"      Total {item['descricao']}")
                apft(c, size=10); aln(c, h="left"); c.border = BORDA_FINA
                val_cell(linha, 2, valor, fundo=fundo)
                
                cp = ws.cell(row=linha, column=3, value=f"{pct:.1f}%" if pct else "—")
                apft(cp, size=10); aln(cp, h="center"); cp.border = BORDA_FINA; apf(cp, fundo)
                
                ws.row_dimensions[linha].height = 18
                linha += 1

        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 20
        
        # Salvar em bytes
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()
