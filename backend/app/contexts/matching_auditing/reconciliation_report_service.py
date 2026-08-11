import uuid
from typing import Dict, Any, List
from datetime import date
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.financeiro import TituloFinanceiro, MovimentacaoFinanceira, StatusTitulo
from app.models.domain import Conciliacao, ConciliacaoItem, StatusConciliacao

class ReconciliationReportService:
    def __init__(self, db: Session):
        self.db = db

    def _to_uuid(self, val):
        if not val: return None
        if isinstance(val, uuid.UUID): return val
        try:
            return uuid.UUID(str(val))
        except Exception:
            return None

    def gerar_relatorio_resumo(self, empresa_id: Any) -> Dict[str, Any]:
        emp_uuid = self._to_uuid(empresa_id)

        # 1. Títulos Efetivados (Liquidados e Conciliados)
        titulos_efetivados = self.db.query(TituloFinanceiro).filter(
            TituloFinanceiro.empresa_id == emp_uuid,
            TituloFinanceiro.status == StatusTitulo.LIQUIDADO
        ).all()

        # 2. Títulos em Aberto (ERP)
        titulos_em_aberto = self.db.query(TituloFinanceiro).filter(
            TituloFinanceiro.empresa_id == emp_uuid,
            TituloFinanceiro.status != StatusTitulo.LIQUIDADO
        ).all()

        # 3. Movimentações Bancárias do Extrato
        movs = self.db.query(MovimentacaoFinanceira).filter(
            MovimentacaoFinanceira.empresa_id == emp_uuid
        ).all()

        movs_conciliadas = [m for m in movs if m.conciliada]
        movs_pendentes = [m for m in movs if not m.conciliada]

        total_pago_efetivado = sum(float(t.valor_pago or t.valor_nominal or 0.0) for t in titulos_efetivados)
        total_em_aberto = sum(float(t.valor_nominal or 0.0) for t in titulos_em_aberto)
        total_saidas_banco = sum(abs(float(m.valor or 0.0)) for m in movs if float(m.valor or 0.0) < 0)
        total_entradas_banco = sum(float(m.valor or 0.0) for m in movs if float(m.valor or 0.0) > 0)

        return {
            "resumo": {
                "total_titulos_erp": len(titulos_efetivados) + len(titulos_em_aberto),
                "titulos_efetivados_count": len(titulos_efetivados),
                "titulos_em_aberto_count": len(titulos_em_aberto),
                "total_pago_efetivado": total_pago_efetivado,
                "total_em_aberto": total_em_aberto,
                "total_movimentos_banco": len(movs),
                "movimentos_conciliados_count": len(movs_conciliadas),
                "movimentos_pendentes_count": len(movs_pendentes),
                "total_saidas_banco": total_saidas_banco,
                "total_entradas_banco": total_entradas_banco,
                "taxa_conciliacao": round((len(titulos_efetivados) / max(len(titulos_efetivados) + len(titulos_em_aberto), 1)) * 100.0, 1)
            },
            "titulos_efetivados": [
                {
                    "descricao": t.descricao,
                    "fornecedor": t.fornecedor_cliente_nome or "",
                    "categoria": t.categoria or "Não Categoria",
                    "valor_nominal": float(t.valor_nominal or 0.0),
                    "valor_pago": float(t.valor_pago or t.valor_nominal or 0.0),
                    "vencimento": t.data_vencimento.strftime("%d/%m/%Y") if t.data_vencimento else "",
                    "pagamento": t.data_pagamento.strftime("%d/%m/%Y") if t.data_pagamento else "",
                    "status": "LIQUIDADO"
                } for t in titulos_efetivados
            ],
            "titulos_em_aberto": [
                {
                    "descricao": t.descricao,
                    "fornecedor": t.fornecedor_cliente_nome or "",
                    "categoria": t.categoria or "Não Categoria",
                    "valor_nominal": float(t.valor_nominal or 0.0),
                    "vencimento": t.data_vencimento.strftime("%d/%m/%Y") if t.data_vencimento else "",
                    "status": "EM ABERTO"
                } for t in titulos_em_aberto
            ]
        }

    def exportar_excel(self, dados: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook()
        
        # Aba 1: Resumo Executivo
        ws_summary = wb.active
        ws_summary.title = "Resumo Conciliação"
        ws_summary.sheet_view.showGridLines = True

        COR_NAVY = "1B2A4A"
        COR_HEADER = "2C5F8A"
        COR_PAR = "F0F4F8"

        font_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=10, bold=True)
        font_data = Font(name="Arial", size=10)

        fill_titulo = PatternFill("solid", fgColor=COR_NAVY)
        fill_header = PatternFill("solid", fgColor=COR_HEADER)
        fill_par = PatternFill("solid", fgColor=COR_PAR)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        border_thin = Border(
            left=Side(style="thin", color="B0BEC5"), right=Side(style="thin", color="B0BEC5"),
            top=Side(style="thin", color="B0BEC5"), bottom=Side(style="thin", color="B0BEC5")
        )

        ws_summary.merge_cells("A1:D1")
        ws_summary["A1"] = "RELATÓRIO DE CONCILIAÇÃO BANCÁRIA & AUDITORIA"
        ws_summary["A1"].font = font_titulo
        ws_summary["A1"].fill = fill_titulo
        ws_summary["A1"].alignment = align_center

        res = dados.get("resumo", {})
        kpis = [
            ("Total de Títulos ERP", res.get("total_titulos_erp", 0)),
            ("Títulos Liquidados e Efetivados", res.get("titulos_efetivados_count", 0)),
            ("Títulos em Aberto (Pendente Banco)", res.get("titulos_em_aberto_count", 0)),
            ("Taxa de Conciliação Bancária", f"{res.get('taxa_conciliacao', 0.0)}%"),
            ("Total Pago Efetivado (R$)", res.get("total_pago_efetivado", 0.0)),
            ("Total em Aberto ERP (R$)", res.get("total_em_aberto", 0.0)),
            ("Total de Saídas no Banco (R$)", res.get("total_saidas_banco", 0.0)),
            ("Total de Entradas no Banco (R$)", res.get("total_entradas_banco", 0.0)),
        ]

        ws_summary.cell(row=3, column=1, value="Indicador / Métrica").font = font_header
        ws_summary.cell(row=3, column=1).fill = fill_header
        ws_summary.cell(row=3, column=2, value="Valor / Quantidade").font = font_header
        ws_summary.cell(row=3, column=2).fill = fill_header

        row_idx = 4
        for label, val in kpis:
            c1 = ws_summary.cell(row=row_idx, column=1, value=label)
            c2 = ws_summary.cell(row=row_idx, column=2, value=val)
            c1.font = font_data
            c2.font = font_bold
            c1.border = border_thin
            c2.border = border_thin
            if isinstance(val, float): c2.number_format = "R$ #,##0.00"
            row_idx += 1

        # Aba 2: Títulos Efetivados
        ws_efetivados = wb.create_sheet(title="Títulos Efetivados")
        ws_efetivados.sheet_view.showGridLines = True
        headers_ef = ["Descrição", "Fornecedor", "Categoria", "Valor Nominal", "Valor Pago", "Vencimento", "Pagamento", "Status"]
        
        for col_num, h in enumerate(headers_ef, 1):
            cell = ws_efetivados.cell(row=1, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        for r_i, item in enumerate(dados.get("titulos_efetivados", []), 2):
            ws_efetivados.cell(row=r_i, column=1, value=item["descricao"])
            ws_efetivados.cell(row=r_i, column=2, value=item["fornecedor"])
            ws_efetivados.cell(row=r_i, column=3, value=item["categoria"])
            c4 = ws_efetivados.cell(row=r_i, column=4, value=item["valor_nominal"])
            c5 = ws_efetivados.cell(row=r_i, column=5, value=item["valor_pago"])
            ws_efetivados.cell(row=r_i, column=6, value=item["vencimento"])
            ws_efetivados.cell(row=r_i, column=7, value=item["pagamento"])
            ws_efetivados.cell(row=r_i, column=8, value=item["status"])
            c4.number_format = "R$ #,##0.00"
            c5.number_format = "R$ #,##0.00"

        # Aba 3: Títulos em Aberto
        ws_aberto = wb.create_sheet(title="Títulos Em Aberto")
        ws_aberto.sheet_view.showGridLines = True
        headers_ab = ["Descrição", "Fornecedor", "Categoria", "Valor Nominal", "Vencimento", "Status"]
        
        for col_num, h in enumerate(headers_ab, 1):
            cell = ws_aberto.cell(row=1, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        for r_i, item in enumerate(dados.get("titulos_em_aberto", []), 2):
            ws_aberto.cell(row=r_i, column=1, value=item["descricao"])
            ws_aberto.cell(row=r_i, column=2, value=item["fornecedor"])
            ws_aberto.cell(row=r_i, column=3, value=item["categoria"])
            c4 = ws_aberto.cell(row=r_i, column=4, value=item["valor_nominal"])
            ws_aberto.cell(row=r_i, column=5, value=item["vencimento"])
            ws_aberto.cell(row=r_i, column=6, value=item["status"])
            c4.number_format = "R$ #,##0.00"

        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        import io
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
