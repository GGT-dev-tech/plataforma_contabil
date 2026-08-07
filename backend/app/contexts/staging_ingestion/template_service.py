import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_standard_template() -> bytes:
    """
    Gera a Planilha Padrão do Sistema em formato .xlsx com 4 abas padronizadas:
    1. Receitas
    2. Despesas
    3. Extrato Bancario
    4. Movimentacao Dinheiro
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Cores e estilos do Design System
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark slate
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    sheets_config = [
        {
            "name": "1_Receitas",
            "headers": ["Data (AAAA-MM-DD)", "Cliente", "Descrição", "Valor R$", "Forma Pagamento", "Conta Destino"],
            "sample_rows": [
                ["2026-06-05", "ACME LOGISTICA LTDA", "Venda de Serviços de Consultoria", 12500.00, "PIX", "Banco Inter"],
                ["2026-06-12", "TECH SOLUTIONS S/A", "Desenvolvimento de Software", 8400.00, "TED", "Banco Inter"]
            ]
        },
        {
            "name": "2_Despesas",
            "headers": ["Data (AAAA-MM-DD)", "Fornecedor", "CNPJ/CPF", "Descrição / Categoria", "Valor R$", "Conta Origem"],
            "sample_rows": [
                ["2026-06-10", "POSTO SHELL PETRO", "12.345.678/0001-99", "Combustível Frota", 350.00, "Banco Inter"],
                ["2026-06-15", "PAPELARIA CENTRAL", "98.765.432/0001-11", "Material de Escritório", 120.50, "Caixa Geral (Dinheiro)"]
            ]
        },
        {
            "name": "3_Extrato_Bancario",
            "headers": ["Data (AAAA-MM-DD)", "Histórico / Descrição", "Valor R$", "Tipo (D/C)", "Banco"],
            "sample_rows": [
                ["2026-06-05", "PIX RECEBIDO ACME LOGISTICA", 12500.00, "C", "Banco Inter"],
                ["2026-06-10", "PGTO PIX POSTO SHELL", -350.00, "D", "Banco Inter"]
            ]
        },
        {
            "name": "4_Movimentacao_Dinheiro",
            "headers": ["Data (AAAA-MM-DD)", "Descrição", "Valor R$", "Tipo (Entrada/Saida)", "Responsavel"],
            "sample_rows": [
                ["2026-06-15", "Pagamento Material Escritório", 120.50, "Saida", "Carlos Almoxarifado"],
                ["2026-06-20", "Suprimento de Caixa para Troco", 500.00, "Entrada", "Maria Financeiro"]
            ]
        }
    ]

    for cfg in sheets_config:
        ws = wb.create_sheet(title=cfg["name"])
        
        # Add Headers
        ws.append(cfg["headers"])
        for col_num in range(1, len(cfg["headers"]) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add Sample Rows
        for row_data in cfg["sample_rows"]:
            ws.append(row_data)

        # Style Rows and set Auto Column Width
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cfg["headers"])):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
