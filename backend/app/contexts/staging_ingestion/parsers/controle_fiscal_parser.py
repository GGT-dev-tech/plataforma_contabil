import pandas as pd
from typing import BinaryIO
from sqlalchemy.orm import Session
import uuid
import openpyxl

from app.models.domain import StagingRegistro, TipoStaging, TipoArquivo
from app.contexts.staging_ingestion.parsers.base import ImportAdapter

class ControleFiscalParser(ImportAdapter):
    """
    Parser para o arquivo padrão controlefiscalcedipi.xlsx
    Aba alvo: "Folha 1"
    """

    def can_parse(self, file_path: str, tipo_arquivo: TipoArquivo) -> bool:
        if tipo_arquivo != TipoArquivo.DESPESA and tipo_arquivo != TipoArquivo.EXTRATO:
            # We accept it as EXTRATO or DESPESA for now
            return False
            
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            return "Folha 1" in wb.sheetnames
        except Exception:
            return False

    def _extrair_categoria(self, resumo: str) -> str:
        if not resumo:
            return "SEM_CATEGORIA"
        linhas = str(resumo).split("\n")
        if len(linhas) >= 2:
            cat = linhas[1].strip()
            return cat if cat else linhas[0].strip()
        return linhas[0].strip()

    def parse(self, file_path: str, db_session: Session, execucao_id: str) -> bool:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["Folha 1"]
        
        registros = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= 2 or not any(row):
                continue
                
            cols = list(row) + [None] * 5
            data_raw, resumo, situacao, valor, _ = cols[:5]
            
            if valor is None or not isinstance(valor, (int, float)):
                continue

            dt = self._parse_date(data_raw)
            if not dt:
                continue
                
            valor_float = float(valor)
            cat = self._extrair_categoria(resumo)
            
            tipo = TipoStaging.RECEITA if valor_float >= 0 else TipoStaging.DESPESA
            
            obs = f"Situação original: {situacao}"
            if str(situacao).strip() not in {"Pago", "Recebido"}:
                obs += " [NÃO EFETIVADO]"

            reg = StagingRegistro(
                id=str(uuid.uuid4()),
                execucao_id=execucao_id,
                tipo=tipo,
                data=dt,
                descricao=str(resumo or ""),
                valor=valor_float,
                categoria=cat,
                observacoes=obs
            )
            registros.append(reg)
            
        if registros:
            db_session.add_all(registros)
            db_session.flush()
            
        return True
