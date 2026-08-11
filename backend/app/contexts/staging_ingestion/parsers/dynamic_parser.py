import openpyxl
import csv
import io
import hashlib
from decimal import Decimal
from datetime import datetime, date
from typing import List, Dict, Any, Tuple
from app.models.domain import TipoStaging

class DynamicTemplateParser:
    """
    Parser Flexível que lê planilhas (XLSX ou CSV) baseado em um de-para (ClientSchemaMapping).
    """

    def parse(self, file_stream, import_config: Dict[str, Any], mapping_json: Dict[str, str] = None, tipo: TipoStaging = TipoStaging.DESPESA) -> List[Dict[str, Any]]:
        """
        Realiza o parse dos dados usando import_config (regras estruturais, ex: skip_rows, col_data)
        e mapping_json (de-para de categorias contábeis).
        """
        # Leitura inicial - tentamos openpyxl (xlsx), senão fallback para CSV
        file_bytes = file_stream.read()
        file_stream.seek(0)
        
        rows = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
        except Exception:
            # Fallback CSV
            text = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(text), delimiter=';') # assumindo ;
            for row in reader:
                rows.append(row)

        if not rows:
            return []

        # Aplica skip_rows
        skip_rows = int(import_config.get("skip_rows", 0)) if import_config else 0
        if len(rows) > skip_rows:
            headers = rows[skip_rows]
            data_rows = rows[skip_rows+1:]
        else:
            return []

        # Normaliza cabecalho
        headers = [h.strip().lower() for h in headers]
        
        # Identifica indices das colunas baseado no config
        if not import_config: import_config = {}
        col_data_name = import_config.get("col_data", "").lower()
        col_valor_name = import_config.get("col_valor", "").lower()
        col_desc_name = import_config.get("col_descricao", "").lower()
        col_entidade_name = import_config.get("col_entidade", "").lower()
        col_categoria_name = import_config.get("col_categoria", "categoria").lower()

        idx_data = headers.index(col_data_name) if col_data_name in headers else -1
        idx_valor = headers.index(col_valor_name) if col_valor_name in headers else -1
        idx_desc = headers.index(col_desc_name) if col_desc_name in headers else -1
        idx_entidade = headers.index(col_entidade_name) if col_entidade_name in headers else -1
        idx_cat = headers.index(col_categoria_name) if col_categoria_name in headers else -1

        def parse_date(val):
            val = str(val).strip()
            if not val: return datetime.utcnow().date().isoformat()
            # Tentar parse ISO ou BR
            if len(val) >= 10:
                if val[4] == '-': return val[:10]
                if val[2] == '/':
                    try:
                        return datetime.strptime(val[:10], "%d/%m/%Y").date().isoformat()
                    except: pass
            return datetime.utcnow().date().isoformat()

        def parse_decimal(val):
            val_str = str(val).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
            try:
                return abs(Decimal(val_str))
            except Exception:
                return Decimal("0.00")

        staging_items = []
        for r in data_rows:
            if not r or (idx_data != -1 and not r[idx_data]):
                continue
                
            # Extração
            raw_data = r[idx_data] if idx_data != -1 and idx_data < len(r) else ""
            raw_valor = r[idx_valor] if idx_valor != -1 and idx_valor < len(r) else "0"
            raw_desc = r[idx_desc] if idx_desc != -1 and idx_desc < len(r) else "Importação Dinâmica"
            raw_entidade = r[idx_entidade] if idx_entidade != -1 and idx_entidade < len(r) else ""
            raw_cat = r[idx_cat] if idx_cat != -1 and idx_cat < len(r) else ""

            # Aplica De-Para de Categoria (se houver mapping_json)
            categoria_final = raw_cat
            if mapping_json and raw_cat in mapping_json:
                categoria_final = mapping_json[raw_cat]

            staging_items.append({
                "tipo": tipo,
                "data": parse_date(raw_data),
                "descricao": raw_desc,
                "valor": parse_decimal(raw_valor),
                "entidade_nome": raw_entidade,
                "categoria": categoria_final,
            })

        return staging_items

    def extract_headers(self, file_stream, skip_rows: int = 0) -> Tuple[str, List[str]]:
        """
        Retorna a assinatura (hash) do cabeçalho e a lista de colunas.
        Usado no frontend para montar o De-Para de colunas caso o ClientSchemaMapping nao exista.
        """
        file_bytes = file_stream.read()
        file_stream.seek(0)
        
        rows = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
        except Exception:
            text = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(text), delimiter=';')
            for row in reader:
                rows.append(row)

        if len(rows) > skip_rows:
            headers = [str(h).strip() for h in rows[skip_rows] if h]
            signature = hashlib.sha256(",".join(headers).encode('utf-8')).hexdigest()
            return signature, headers
        return "", []
