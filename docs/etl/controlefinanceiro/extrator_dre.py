"""
=======================================================================
EXTRATOR DRE / BALANCETE  v2.0
=======================================================================
Autor   : Gerado via Antigravity (Google DeepMind)
Versão  : 2.0.0

COMO USAR:
  1. Coloque os arquivos mensais em:
         controlefinanceiro/entrada/
     com o nome:
         controlefinanceiro_MMAAAA.xlsx
     Exemplo: controlefinanceiro_012026.xlsx (janeiro/2026)
              controlefinanceiro_072026.xlsx (julho/2026)

  2. Execute no terminal:
         python3 extrator_dre.py

  3. Os DREs são gerados em:
         controlefinanceiro/saida/AAAA/DRE_Balancete_MMAAAA.xlsx

  4. Arquivos originais são copiados automaticamente para:
         controlefinanceiro/backup/

  5. O histórico acumulado fica em:
         controlefinanceiro/historico.json

MESES JÁ PROCESSADOS são ignorados automaticamente.
Para reprocessar um mês, remova sua chave do historico.json.

Dependência: pip install openpyxl
=======================================================================
"""

import os
import sys
import json
import re
import shutil
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Dependencia faltando. Execute:  pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────
# 1. CONFIGURAÇÃO  ← edite aqui sem mexer no restante do código
# ─────────────────────────────────────────────────────────────────────

ABA_ENTRADA          = "Folha 1"
DIR_ENTRADA          = Path("entrada")
DIR_SAIDA            = Path("saida")
DIR_BACKUP           = Path("backup")
HIST_PATH            = Path("historico.json")
SITUACOES_EFETIVADAS = {"Pago", "Recebido"}
PADRAO_ARQUIVO       = re.compile(
    r"(?:controlefinanceiro|controlefiscalcedipi)_(\d{2})(\d{4})\.xlsx$", re.IGNORECASE
)

MESES_PT = {
    1: "Janeiro",  2: "Fevereiro", 3: "Marco",    4: "Abril",
    5: "Maio",     6: "Junho",     7: "Julho",     8: "Agosto",
    9: "Setembro", 10: "Outubro",  11: "Novembro", 12: "Dezembro",
}

# ── Mapeamento  categoria → grupo DRE ──────────────────────────────
MAPA_DRE = {
    "Receitas de Vendas"                          : "RECEITA_BRUTA",
    "Receitas de Serviços"                        : "RECEITA_BRUTA",
    "Cartão Maquininha Recebimento"               : "RECEITA_BRUTA",
    "PMBC (Prefeitura de Balneário Camboriú)"     : "RECEITA_BRUTA",
    "Impostos"                                    : "DEDUCOES_RECEITA",
    "Impostos sobre Serviços (ISS)"               : "DEDUCOES_RECEITA",
    "Materiais Aplicados na Prestação de Serviços": "CSP",
    "Agulhas biópsias"                            : "CSP",
    "Honorários Médicos"                          : "CSP",
    "Dr. Wagner (Honorários)"                     : "CSP",
    "Dr. Deison (Honorários)"                     : "CSP",
    "Salários"                                    : "DESPESAS_PESSOAL",
    "INSS sobre Salários - GPS"                   : "DESPESAS_PESSOAL",
    "FGTS e Multa de FGTS"                        : "DESPESAS_PESSOAL",
    "Vale-Alimentação"                            : "DESPESAS_PESSOAL",
    "Uniformes"                                   : "DESPESAS_PESSOAL",
    "Honorários Contábeis"                        : "DESPESAS_ADMINISTRATIVAS",
    "Honorários Consultoria"                      : "DESPESAS_ADMINISTRATIVAS",
    "certificadora"                               : "DESPESAS_ADMINISTRATIVAS",
    "Bens de Pequeno Valor"                       : "DESPESAS_ADMINISTRATIVAS",
    "Computadores e Periféricos"                  : "DESPESAS_ADMINISTRATIVAS",
    "Aluguel"                                     : "DESPESAS_INSTALACOES",
    "Energia Elétrica"                            : "DESPESAS_INSTALACOES",
    "Água e Saneamento"                           : "DESPESAS_INSTALACOES",
    "Agua Bombona"                                : "DESPESAS_INSTALACOES",
    "Telefonia e Internet"                        : "DESPESAS_INSTALACOES",
    "Telefonia Móvel"                             : "DESPESAS_INSTALACOES",
    "Vigilância e Segurança Patrimonial"          : "DESPESAS_INSTALACOES",
    "Manutenção Predial"                          : "DESPESAS_INSTALACOES",
    "Split Manutenção"                            : "DESPESAS_INSTALACOES",
    "Materiais de Limpeza e de Higiene"           : "DESPESAS_INSTALACOES",
    "Marketing"                                   : "DESPESAS_COMERCIAIS",
    "Google"                                      : "DESPESAS_COMERCIAIS",
    "Brindes para Clientes"                       : "DESPESAS_COMERCIAIS",
    "Descontos financeiros obtidos"               : "OUTRAS_RECEITAS",
    "Antecipação de Lucros"                       : "DISTRIBUICAO_LUCROS",
    "antecipação de lucros - Honorários"          : "CSP",
    "Taxas Bancárias"                             : "DESPESAS_ADMINISTRATIVAS",
    "Tarifas"                                     : "DESPESAS_ADMINISTRATIVAS",
    "Material de Escritório"                      : "DESPESAS_ADMINISTRATIVAS",
    "Combustível"                                 : "DESPESAS_COMERCIAIS",
    "Estacionamento"                              : "DESPESAS_COMERCIAIS",
    "Viagens"                                     : "DESPESAS_COMERCIAIS",
    "Refeições"                                   : "DESPESAS_COMERCIAIS",
    "Software"                                    : "DESPESAS_ADMINISTRATIVAS",
    "Assinaturas"                                 : "DESPESAS_ADMINISTRATIVAS",
    "Treinamentos"                                : "DESPESAS_PESSOAL",
    "Uniforme Equipe"                             : "DESPESAS_PESSOAL",
    "Manutenção Equipamentos"                     : "DESPESAS_INSTALACOES",
    "Reposição Materiais"                         : "CSP",
    "Exames Externos"                             : "CSP",
    "Descarte Resíduos"                           : "DESPESAS_INSTALACOES",
    "Seguros"                                     : "DESPESAS_ADMINISTRATIVAS",
    "Licenças"                                    : "DESPESAS_ADMINISTRATIVAS",
    "Certificados Digitais"                       : "DESPESAS_ADMINISTRATIVAS",
    "Contribuições Sindicais"                     : "DESPESAS_PESSOAL",
    "Benefícios"                                  : "DESPESAS_PESSOAL",
    "Confraternizações"                           : "DESPESAS_PESSOAL",
    "Juros Pagos"                                 : "DESPESAS_ADMINISTRATIVAS",
    "Multas Pagas"                                : "DESPESAS_ADMINISTRATIVAS",
    "Correios"                                    : "DESPESAS_ADMINISTRATIVAS",
    "Limpeza e Conservação"                       : "DESPESAS_INSTALACOES",
    "Reparos Gerais"                              : "DESPESAS_INSTALACOES",
    "Equipamentos Médicos"                        : "DESPESAS_ADMINISTRATIVAS",
    "Cursos de Aperfeiçoamento"                   : "DESPESAS_PESSOAL",
    "Palestras"                                   : "DESPESAS_COMERCIAIS",
    "Eventos"                                     : "DESPESAS_COMERCIAIS",
    "Taxas Públicas"                              : "DEDUCOES_RECEITA",
    "Outras Despesas Administrativas"             : "DESPESAS_ADMINISTRATIVAS",

    # ── Categorias exatas do arquivo controlefiscalcedipi (Jan–Jul) ─

    # Deduções / Tributos sobre Receita
    "IRRF Darf 567"                               : "DEDUCOES_RECEITA",
    "Retenção - ISS Serviços Tomados"             : "DEDUCOES_RECEITA",

    # Custo dos Serviços Prestados (variação de grafia)
    "Agulhas Biópsias"                            : "CSP",

    # Despesas com Pessoal
    "Adiantamento Salarial"                       : "DESPESAS_PESSOAL",
    "Férias"                                      : "DESPESAS_PESSOAL",
    "Rescisões"                                   : "DESPESAS_PESSOAL",
    "Remuneração de Estagiários"                  : "DESPESAS_PESSOAL",
    "Seguro de Vida"                              : "DESPESAS_PESSOAL",
    "Exames Médicos"                              : "DESPESAS_PESSOAL",
    "reembolso"                                   : "DESPESAS_PESSOAL",
    "reembolso de APP transporte"                 : "DESPESAS_PESSOAL",

    # Despesas Administrativas
    "Honorários Advocatícios"                     : "DESPESAS_ADMINISTRATIVAS",
    "Alvara Sanitário"                            : "DESPESAS_ADMINISTRATIVAS",
    "Alvará de Funcionamento"                     : "DESPESAS_ADMINISTRATIVAS",
    "CRM"                                         : "DESPESAS_ADMINISTRATIVAS",
    "Cursos e Treinamentos"                       : "DESPESAS_ADMINISTRATIVAS",
    "ACORDO JUDICIAL"                             : "DESPESAS_ADMINISTRATIVAS",
    "PASTAS E FOLHAS"                             : "DESPESAS_ADMINISTRATIVAS",
    "Materiais de Escritório"                     : "DESPESAS_ADMINISTRATIVAS",

    # Despesas com Instalações
    "IPTU"                                        : "DESPESAS_INSTALACOES",
    "Taxa de Lixo"                                : "DESPESAS_INSTALACOES",
    "Manutenção de Equipamentos"                  : "DESPESAS_INSTALACOES",
    "Material de limpeza + bombona"               : "DESPESAS_INSTALACOES",

    # Despesas Comerciais
    "Marketing e Publicidade"                     : "DESPESAS_COMERCIAIS",
    "google"                                      : "DESPESAS_COMERCIAIS",

    # Outras Receitas (entradas não operacionais)
    "Empréstimos de Sócios"                       : "OUTRAS_RECEITAS",
    "Rendimentos de Aplicações"                   : "OUTRAS_RECEITAS",

    # NAO_MAPEADO intencional — não entram no DRE (vão para alerta)
    # "transferencia entre contas" → movimento interno entre contas, neutro
    # "Despesas a identificar"    → sem classificação contábil segura
}

ROTULOS_GRUPO = {
    "RECEITA_BRUTA"            : "Receita Bruta",
    "DEDUCOES_RECEITA"         : "Deduções da Receita (Tributos)",
    "CSP"                      : "Custo dos Serviços Prestados",
    "DESPESAS_PESSOAL"         : "Despesas com Pessoal",
    "DESPESAS_ADMINISTRATIVAS" : "Despesas Administrativas",
    "DESPESAS_INSTALACOES"     : "Despesas com Instalações",
    "DESPESAS_COMERCIAIS"      : "Despesas Comerciais",
    "OUTRAS_RECEITAS"          : "Outras Receitas",
    "DISTRIBUICAO_LUCROS"      : "Distribuição de Lucros / Pró-labore",
}

ESTRUTURA_DRE = [
    ("RECEITA_BRUTA",            False, None),
    ("DEDUCOES_RECEITA",         True,  "( = ) RECEITA LIQUIDA"),
    ("CSP",                      True,  "( = ) LUCRO BRUTO"),
    ("DESPESAS_PESSOAL",         True,  None),
    ("DESPESAS_ADMINISTRATIVAS", True,  None),
    ("DESPESAS_INSTALACOES",     True,  None),
    ("DESPESAS_COMERCIAIS",      True,  "( = ) RESULTADO OPERACIONAL LIQUIDO"),
    ("OUTRAS_RECEITAS",          False, "( = ) RESULTADO ANTES DA DISTRIBUICAO"),
    ("DISTRIBUICAO_LUCROS",      True,  "( = ) RESULTADO FINAL APOS DISTRIBUICAO"),
]

# ─────────────────────────────────────────────────────────────────────
# 2. PALETA E ESTILOS
# ─────────────────────────────────────────────────────────────────────

COR_TITULO_FUNDO   = "1B2A4A"
COR_TITULO_FONTE   = "FFFFFF"
COR_GRUPO_FUNDO    = "2C5F8A"
COR_GRUPO_FONTE    = "FFFFFF"
COR_LINHA_PAR      = "F0F4F8"
COR_LINHA_IMPAR    = "FFFFFF"
COR_POSITIVO       = "1A6B34"
COR_NEGATIVO       = "8B1A1A"
COR_RESULTADO_POS  = "0D4A24"
COR_RESULTADO_NEG  = "6B0E0E"
COR_HEADER_ABA     = "0A1628"
COR_INFO_FUNDO     = "EBF5FB"   # azul muito claro para linhas informativas
COR_INFO_FONTE     = "1B4F72"
COR_ACUM_FUNDO     = "1A252F"   # cinza escuro para resultado acumulado

BORDA_FINA = Border(
    left=Side(style="thin", color="B0BEC5"),
    right=Side(style="thin", color="B0BEC5"),
    top=Side(style="thin", color="B0BEC5"),
    bottom=Side(style="thin", color="B0BEC5"),
)
BORDA_MEDIA = Border(
    left=Side(style="medium", color="2C5F8A"),
    right=Side(style="medium", color="2C5F8A"),
    top=Side(style="medium", color="2C5F8A"),
    bottom=Side(style="medium", color="2C5F8A"),
)

FMT_MOEDA = '#,##0.00'


# ─────────────────────────────────────────────────────────────────────
# 3. FUNÇÕES DE ESTILO
# ─────────────────────────────────────────────────────────────────────

def apf(cell, hex_cor):
    cell.fill = PatternFill("solid", fgColor=hex_cor)


def apft(cell, bold=False, italic=False, size=10, cor="000000"):
    cell.font = Font(bold=bold, italic=italic, size=size,
                     color=cor, name="Calibri")


def aln(cell, h="left", v="center", wrap=False):
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def hdr(ws, row, colunas, fundo=COR_HEADER_ABA):
    for i, texto in enumerate(colunas, start=1):
        c = ws.cell(row=row, column=i, value=texto)
        apf(c, fundo)
        apft(c, bold=True, cor=COR_TITULO_FONTE, size=10)
        aln(c, h="center")
        c.border = BORDA_FINA
    ws.row_dimensions[row].height = 22


def val_cell(ws, row, col, valor, bold=False, fundo=None,
             cor_fonte=None, total=False):
    v = valor if valor is not None and valor != 0 else None
    c = ws.cell(row=row, column=col, value=v)
    c.number_format = FMT_MOEDA
    if fundo:
        apf(c, fundo)
    if cor_fonte is None:
        cor_fonte = COR_POSITIVO if (valor or 0) >= 0 else COR_NEGATIVO
    apft(c, bold=bold or total, size=10 if not total else 11, cor=cor_fonte)
    aln(c, h="right")
    c.border = BORDA_MEDIA if total else BORDA_FINA
    return c


def grp_cell(ws, row, col, texto, fundo=COR_GRUPO_FUNDO,
             fonte=COR_GRUPO_FONTE):
    c = ws.cell(row=row, column=col, value=texto)
    apf(c, fundo)
    apft(c, bold=True, size=10, cor=fonte)
    aln(c, h="left")
    c.border = BORDA_FINA
    return c


# ─────────────────────────────────────────────────────────────────────
# 4. FUNÇÕES DE INFRAESTRUTURA (NOVO v2)
# ─────────────────────────────────────────────────────────────────────

def criar_estrutura():
    """Cria a estrutura de pastas na primeira execução."""
    DIR_ENTRADA.mkdir(exist_ok=True)
    DIR_SAIDA.mkdir(exist_ok=True)
    DIR_BACKUP.mkdir(exist_ok=True)
    if DIR_ENTRADA.exists() and not any(DIR_ENTRADA.iterdir()):
        print(f"\n  PASTA CRIADA: {DIR_ENTRADA.resolve()}")
        print(f"  -> Coloque seus arquivos mensais aqui com o nome:")
        print(f"     controlefinanceiro_MMAAAA.xlsx")
        print(f"     Exemplo: controlefinanceiro_012026.xlsx")


def ler_historico():
    """Lê o historico.json ou retorna estrutura vazia."""
    if HIST_PATH.exists():
        with open(HIST_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"versao": "2.0", "ultima_atualizacao": None, "meses": {}}


def salvar_historico(hist):
    """Persiste o historico.json."""
    hist["ultima_atualizacao"] = datetime.now().isoformat()
    with open(HIST_PATH, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)


def fazer_backup(caminho: Path, mes: int, ano: int):
    """Copia o arquivo original para backup/ com timestamp."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = f"controlefinanceiro_{mes:02d}{ano}_backup_{ts}.xlsx"
    dest = DIR_BACKUP / nome
    shutil.copy2(caminho, dest)
    return dest


def descobrir_arquivos(hist: dict) -> list:
    """
    Varre entrada/ por arquivos controlefinanceiro_MMAAAA.xlsx.
    Retorna lista ordenada de (mes, ano, Path) ainda não processados.
    """
    if not DIR_ENTRADA.exists() or not any(DIR_ENTRADA.iterdir()):
        # Tenta modo legado: arquivo único na pasta raiz
        f = Path("controlefinanceiro.xlsx")
        if f.exists():
            print(f"\n  Modo compatível (v1): usando {f.name}")
            return [(-1, -1, f)]  # mes=-1 indica modo legado
        return []

    encontrados = []
    for f in DIR_ENTRADA.iterdir():
        m = PADRAO_ARQUIVO.match(f.name)
        if m:
            mes, ano = int(m.group(1)), int(m.group(2))
            chave = f"{ano}-{mes:02d}"
            if chave in hist.get("meses", {}):
                print(f"  [IGNORADO] {f.name} — já processado. "
                      f"Remova '{chave}' do historico.json para reprocessar.")
            else:
                encontrados.append((mes, ano, f))

    return sorted(encontrados, key=lambda x: (x[1], x[0]))


def obter_saldo_final_planilha(caminho: Path, nome_aba: str) -> float:
    """Extrai o último saldo registrado na coluna E da planilha."""
    wb = openpyxl.load_workbook(str(caminho))
    if nome_aba not in wb.sheetnames:
        return 0.0
    ws = wb[nome_aba]
    ultimo = 0.0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[4] is not None and isinstance(row[4], (int, float)):
            ultimo = float(row[4])
    return ultimo


def obter_saldo_anterior(hist: dict, mes: int, ano: int) -> float:
    """Retorna o saldo_final do mês imediatamente anterior registrado."""
    if mes <= 0:
        return 0.0
    mes_ant = mes - 1 if mes > 1 else 12
    ano_ant = ano if mes > 1 else ano - 1
    chave   = f"{ano_ant}-{mes_ant:02d}"
    return hist.get("meses", {}).get(chave, {}).get("saldo_final", 0.0)


def calcular_acumulado_ano(hist: dict, mes: int, ano: int,
                            resultado_mes: float) -> float:
    """Soma os resultados_final de Jan até o mês atual do mesmo ano."""
    acum = resultado_mes
    for m in range(1, mes):
        chave = f"{ano}-{m:02d}"
        dados = hist.get("meses", {}).get(chave, {})
        acum += dados.get("subtotais", {}).get("resultado_final", 0.0)
    return acum


# ─────────────────────────────────────────────────────────────────────
# 5. LEITURA E CLASSIFICAÇÃO
# ─────────────────────────────────────────────────────────────────────

def extrair_categoria(resumo):
    if not resumo:
        return "SEM_CATEGORIA"
    linhas = str(resumo).split("\n")
    if len(linhas) >= 2:
        cat = linhas[1].strip()
        return cat if cat else linhas[0].strip()
    return linhas[0].strip()


def parse_data(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                continue
    return None


def ler_lancamentos(caminho: str, nome_aba: str):
    """
    Lê e classifica todos os lançamentos da planilha.
    Retorna (efetivados, em_aberto, nao_mapeados).
    """
    wb = openpyxl.load_workbook(caminho)
    if nome_aba not in wb.sheetnames:
        raise ValueError(f"Aba '{nome_aba}' não encontrada. "
                         f"Disponíveis: {wb.sheetnames}")
    ws  = wb[nome_aba]
    efe = []
    ema = []
    nao = set()

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 2 or not any(row):
            continue
        cols            = list(row) + [None] * 5
        data_raw, resumo, situacao, valor, _ = cols[:5]
        if valor is None or not isinstance(valor, (int, float)):
            continue

        dt    = parse_data(data_raw)
        cat   = extrair_categoria(resumo)
        grupo = MAPA_DRE.get(cat, "NAO_MAPEADO")
        if grupo == "NAO_MAPEADO":
            nao.add(cat)

        item = {
            "data"      : dt,
            "data_str"  : dt.strftime("%d/%m/%Y") if dt else str(data_raw),
            "resumo"    : str(resumo or ""),
            "categoria" : cat,
            "grupo_dre" : grupo,
            "rotulo_dre": ROTULOS_GRUPO.get(grupo, grupo),
            "situacao"  : str(situacao or ""),
            "valor"     : float(valor),
        }

        if str(situacao) in SITUACOES_EFETIVADAS:
            efe.append(item)
        else:
            ema.append(item)

    return efe, ema, nao


def extrair_resumo_tania(efetivados: list, em_aberto: list) -> dict:
    """Isola os valores recebidos e a receber relacionados a 'Tania'."""
    stats = {"pix": 0.0, "dinheiro": 0.0, "outros": 0.0, "a_receber": 0.0}
    
    for l in efetivados:
        if "tania" in str(l["resumo"]).lower():
            resumo = str(l["resumo"]).lower()
            if "pix" in resumo:
                stats["pix"] += l["valor"]
            elif "dinheiro" in resumo:
                stats["dinheiro"] += l["valor"]
            else:
                stats["outros"] += l["valor"]
                
    for l in em_aberto:
        if "tania" in str(l["resumo"]).lower():
            stats["a_receber"] += l["valor"]
            
    return stats

# ─────────────────────────────────────────────────────────────────────
# 6. CÁLCULO DE SUBTOTAIS (isolado para reuso)
# ─────────────────────────────────────────────────────────────────────

def calcular_subtotais(totais: dict) -> dict:
    """Calcula toda a cascata do DRE e retorna um dict de subtotais."""
    rb  = totais.get("RECEITA_BRUTA", 0.0)
    ded = totais.get("DEDUCOES_RECEITA", 0.0)
    csp = totais.get("CSP", 0.0)
    dp  = totais.get("DESPESAS_PESSOAL", 0.0)
    da  = totais.get("DESPESAS_ADMINISTRATIVAS", 0.0)
    di  = totais.get("DESPESAS_INSTALACOES", 0.0)
    dc  = totais.get("DESPESAS_COMERCIAIS", 0.0)
    or_ = totais.get("OUTRAS_RECEITAS", 0.0)
    dl  = totais.get("DISTRIBUICAO_LUCROS", 0.0)

    rl  = rb  + ded          # receita líquida
    lb  = rl  + csp          # lucro bruto
    ro  = lb  + dp + da + di + dc   # resultado operacional
    rad = ro  + or_          # resultado antes distribuição
    rf  = rad + dl           # resultado final

    return {
        "receita_liquida"              : round(rl, 2),
        "lucro_bruto"                  : round(lb, 2),
        "resultado_operacional"        : round(ro, 2),
        "resultado_antes_distribuicao" : round(rad, 2),
        "resultado_final"              : round(rf, 2),
    }


# ─────────────────────────────────────────────────────────────────────
# 7. GERAÇÃO DRE (atualizado v2 com saldo e acumulado)
# ─────────────────────────────────────────────────────────────────────

def gerar_aba_dre(wb_out, lancamentos, periodo, nao_mapeados,
                  saldo_inicial=0.0, saldo_final=0.0,
                  resultado_acumulado=0.0, mes=None, ano=None,
                  tania_stats=None):
    ws = wb_out.create_sheet("DRE")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    # ── Título ─────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1,
                value=f"DEMONSTRACAO DO RESULTADO DO EXERCICIO — {periodo}")
    apf(c, COR_TITULO_FUNDO); apft(c, bold=True, size=13, cor=COR_TITULO_FONTE)
    aln(c, h="center"); ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    c = ws.cell(row=2, column=1,
                value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} "
                      f"| Regime de Caixa | Efetivados (Pago / Recebido)")
    apf(c, "2C5F8A"); apft(c, italic=True, size=9, cor="E8F4FD")
    aln(c, h="center"); ws.row_dimensions[2].height = 16

    # ── Linha de saldo inicial (informativa) ───────────────────────
    ws.row_dimensions[3].height = 6

    def linha_info(row, texto, valor, cor_valor=COR_INFO_FONTE):
        ws.merge_cells(f"A{row}:A{row}")
        c = ws.cell(row=row, column=1, value=texto)
        apf(c, COR_INFO_FUNDO); apft(c, italic=True, size=9, cor=COR_INFO_FONTE)
        aln(c, h="left"); c.border = BORDA_FINA
        cv = ws.cell(row=row, column=2, value=valor)
        cv.number_format = FMT_MOEDA
        apf(cv, COR_INFO_FUNDO)
        cv.font = Font(italic=True, size=9, color=cor_valor, name="Calibri")
        cv.alignment = Alignment(horizontal="right", vertical="center")
        cv.border = BORDA_FINA
        ce = ws.cell(row=row, column=3, value="(informativo — fora do DRE)")
        apf(ce, COR_INFO_FUNDO); apft(ce, italic=True, size=8, cor="AAB7B8")
        aln(ce, h="center"); ce.border = BORDA_FINA
        ws.row_dimensions[row].height = 16

    linha_info(4, "  Saldo bancario em caixa — inicio do periodo:", saldo_inicial)
    ws.row_dimensions[5].height = 6

    # ── Cabeçalho das colunas ───────────────────────────────────────
    hdr(ws, 6, ["DESCRICAO", "VALOR (R$)", "% RECEITA BRUTA"])
    ws.row_dimensions[7].height = 4

    # ── Totais por grupo ────────────────────────────────────────────
    totais = defaultdict(float)
    for l in lancamentos:
        if l["grupo_dre"] != "NAO_MAPEADO":
            totais[l["grupo_dre"]] += l["valor"]

    rb = totais.get("RECEITA_BRUTA", 0.0)

    def pct(v):
        return round(abs(v) / rb * 100, 2) if rb else None

    linha  = 8
    saldo  = 0.0

    for grupo, subtrativo, label_sub in ESTRUTURA_DRE:
        vg = totais.get(grupo, 0.0)

        # cabeçalho do grupo
        grp_cell(ws, linha, 1, f"  {ROTULOS_GRUPO[grupo]}")
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha, end_column=3)
        ws.row_dimensions[linha].height = 20
        linha += 1

        # valor do grupo
        fundo = COR_LINHA_PAR if linha % 2 == 0 else COR_LINHA_IMPAR
        c = ws.cell(row=linha, column=1,
                    value=f"      Total {ROTULOS_GRUPO[grupo]}")
        apft(c, size=10); aln(c, h="left"); c.border = BORDA_FINA
        val_cell(ws, linha, 2, vg, fundo=fundo)
        p  = pct(vg)
        cp = ws.cell(row=linha, column=3,
                     value=f"{p:.1f}%" if p is not None else "—")
        apft(cp, size=10); aln(cp, h="center"); cp.border = BORDA_FINA
        apf(cp, fundo)
        ws.row_dimensions[linha].height = 18
        linha += 1

        # atualiza saldo cascata
        if grupo == "RECEITA_BRUTA":
            saldo += abs(vg) if vg >= 0 else vg
        elif grupo == "OUTRAS_RECEITAS":
            saldo += vg
        elif subtrativo:
            saldo -= abs(vg)

        if label_sub:
            ws.row_dimensions[linha].height = 6
            linha += 1
            cf = COR_RESULTADO_POS if saldo >= 0 else COR_RESULTADO_NEG
            cd = ws.cell(row=linha, column=1, value=label_sub)
            apf(cd, cf); apft(cd, bold=True, size=11, cor="FFFFFF")
            aln(cd, h="left"); cd.border = BORDA_MEDIA
            val_cell(ws, linha, 2, round(saldo, 2),
                     bold=True, fundo=cf, cor_fonte="FFFFFF", total=True)
            pt  = pct(saldo)
            cpt = ws.cell(row=linha, column=3,
                          value=f"{pt:.1f}%" if pt else "—")
            apf(cpt, cf); apft(cpt, bold=True, cor="FFFFFF", size=11)
            aln(cpt, h="center"); cpt.border = BORDA_MEDIA
            ws.row_dimensions[linha].height = 24
            linha += 2

    # ── Linha saldo final (informativa) ─────────────────────────────
    ws.row_dimensions[linha].height = 6
    linha += 1
    linha_info(linha, "  Saldo bancario em caixa — fim do periodo:", saldo_final)
    linha += 1
    ws.row_dimensions[linha].height = 6
    linha += 1

    # ── Resultado acumulado no ano ───────────────────────────────────
    cf_acum = COR_RESULTADO_POS if resultado_acumulado >= 0 else COR_RESULTADO_NEG
    label_acum = (f"RESULTADO ACUMULADO NO ANO"
                  f"{f' ({MESES_PT[1]} a {MESES_PT[mes]})' if mes else ''}")
    ca = ws.cell(row=linha, column=1, value=label_acum)
    apf(ca, COR_ACUM_FUNDO); apft(ca, bold=True, size=11, cor="FFFFFF")
    aln(ca, h="left"); ca.border = BORDA_MEDIA
    val_cell(ws, linha, 2, round(resultado_acumulado, 2),
             bold=True, fundo=COR_ACUM_FUNDO, cor_fonte="FFFFFF", total=True)
    ce2 = ws.cell(row=linha, column=3, value="")
    apf(ce2, COR_ACUM_FUNDO); ce2.border = BORDA_MEDIA
    ws.row_dimensions[linha].height = 26
    linha += 2

    # ── Alertas categorias não mapeadas ─────────────────────────────
    if nao_mapeados:
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha, end_column=3)
        c = ws.cell(row=linha, column=1,
                    value="ATENCAO: Categorias nao mapeadas "
                          "(verifique MAPA_DRE no script):")
        apf(c, "FFF3CD"); apft(c, bold=True, cor="856404", size=10)
        aln(c, h="left"); linha += 1
        for cat in sorted(nao_mapeados):
            ws.merge_cells(start_row=linha, start_column=1,
                           end_row=linha, end_column=3)
            c = ws.cell(row=linha, column=1, value=f"     * {cat}")
            apf(c, "FFF8E1"); apft(c, cor="856404", size=9)
            aln(c, h="left"); linha += 1

    # ── Resumo Recebimentos Tania ───────────────────────────────────
    if tania_stats and any(tania_stats.values()):
        ws.row_dimensions[linha].height = 10
        linha += 1
        
        ca = ws.cell(row=linha, column=1, value="  RESUMO DE RECEBIMENTOS - TANIA")
        apf(ca, COR_GRUPO_FUNDO); apft(ca, bold=True, size=11, cor="FFFFFF")
        aln(ca, h="left"); ca.border = BORDA_MEDIA
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
        ws.row_dimensions[linha].height = 20
        linha += 1
        
        def linha_tania(titulo, valor, cor_fundo, cor_texto="000000"):
            c_tit = ws.cell(row=linha, column=1, value=f"      {titulo}")
            apf(c_tit, cor_fundo); apft(c_tit, cor=cor_texto); c_tit.border = BORDA_FINA
            
            c_val = ws.cell(row=linha, column=2, value=valor)
            c_val.number_format = FMT_MOEDA
            apf(c_val, cor_fundo); apft(c_val, cor=cor_texto, bold=True if "A Receber" in titulo else False)
            aln(c_val, h="right"); c_val.border = BORDA_FINA
            
            c_emp = ws.cell(row=linha, column=3, value="")
            apf(c_emp, cor_fundo); c_emp.border = BORDA_FINA
            
        linha_tania("Recebido via PIX", tania_stats["pix"], COR_LINHA_IMPAR)
        linha += 1
        linha_tania("Recebido em Dinheiro", tania_stats["dinheiro"], COR_LINHA_PAR)
        linha += 1
        linha_tania("Recebido (Outras Vendas/TED)", tania_stats["outros"], COR_LINHA_IMPAR)
        linha += 1
        linha_tania("Em Aberto (A Receber)", tania_stats["a_receber"], "FFF8E1", cor_texto="856404")
        linha += 1
        ws.row_dimensions[linha].height = 6
        linha += 1

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20


# ─────────────────────────────────────────────────────────────────────
# 8. GERAÇÃO BALANCETE ANALÍTICO (igual v1)
# ─────────────────────────────────────────────────────────────────────

def gerar_aba_balancete(wb_out, lancamentos, periodo):
    ws = wb_out.create_sheet("Balancete Analitico")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1,
                value=f"BALANCETE ANALITICO — {periodo}")
    apf(c, COR_TITULO_FUNDO); apft(c, bold=True, size=13, cor=COR_TITULO_FONTE)
    aln(c, h="center"); ws.row_dimensions[1].height = 28

    hdr(ws, 2, ["GRUPO DRE", "CATEGORIA", "QTD",
                "TOTAL DEBITOS (R$)", "TOTAL CREDITOS (R$)"])

    estrutura = defaultdict(lambda: defaultdict(list))
    for l in lancamentos:
        estrutura[l["grupo_dre"]][l["categoria"]].append(l["valor"])

    ORDEM  = [g for g, _, _ in ESTRUTURA_DRE]
    extras = [g for g in estrutura if g not in ORDEM]

    linha          = 3
    tot_geral_deb  = 0.0
    tot_geral_cred = 0.0

    for grupo in ORDEM + extras:
        if grupo not in estrutura:
            continue
        rotulo = ROTULOS_GRUPO.get(grupo, grupo)
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha, end_column=5)
        grp_cell(ws, linha, 1, f"  {rotulo}")
        ws.row_dimensions[linha].height = 20
        linha += 1

        tot_deb = tot_cred = 0.0
        for cat, vals in sorted(estrutura[grupo].items()):
            deb  = sum(v for v in vals if v < 0)
            cred = sum(v for v in vals if v >= 0)
            tot_deb  += deb
            tot_cred += cred
            fundo = COR_LINHA_PAR if linha % 2 == 0 else COR_LINHA_IMPAR

            def fc(col, val, h="left"):
                c = ws.cell(row=linha, column=col, value=val)
                apf(c, fundo); apft(c, size=9); aln(c, h=h); c.border = BORDA_FINA
                return c

            fc(1, ""); fc(2, cat); fc(3, len(vals), h="center")
            val_cell(ws, linha, 4, deb if deb else None, fundo=fundo)
            val_cell(ws, linha, 5, cred if cred else None, fundo=fundo)
            ws.row_dimensions[linha].height = 16
            linha += 1

        cs = ws.cell(row=linha, column=1, value=f"Subtotal — {rotulo}")
        apf(cs, COR_GRUPO_FUNDO); apft(cs, bold=True, cor="FFFFFF", size=9)
        aln(cs, h="left"); cs.border = BORDA_FINA
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha, end_column=2)
        c3 = ws.cell(row=linha, column=3, value="")
        apf(c3, COR_GRUPO_FUNDO); c3.border = BORDA_FINA
        val_cell(ws, linha, 4, tot_deb if tot_deb else None,
                 bold=True, fundo=COR_GRUPO_FUNDO, cor_fonte="FFFFFF")
        val_cell(ws, linha, 5, tot_cred if tot_cred else None,
                 bold=True, fundo=COR_GRUPO_FUNDO, cor_fonte="FFFFFF")
        tot_geral_deb  += tot_deb
        tot_geral_cred += tot_cred
        ws.row_dimensions[linha].height = 18
        linha += 2

    ctg = ws.cell(row=linha, column=1, value="TOTAL GERAL")
    apf(ctg, COR_TITULO_FUNDO); apft(ctg, bold=True, size=11, cor="FFFFFF")
    aln(ctg, h="left"); ctg.border = BORDA_MEDIA
    ws.merge_cells(start_row=linha, start_column=1,
                   end_row=linha, end_column=3)
    val_cell(ws, linha, 4, tot_geral_deb, bold=True,
             fundo=COR_TITULO_FUNDO, cor_fonte="FFFFFF", total=True)
    val_cell(ws, linha, 5, tot_geral_cred, bold=True,
             fundo=COR_TITULO_FUNDO, cor_fonte="FFFFFF", total=True)
    ws.row_dimensions[linha].height = 24

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20


# ─────────────────────────────────────────────────────────────────────
# 9. GERAÇÃO LANÇAMENTOS (igual v1)
# ─────────────────────────────────────────────────────────────────────

def gerar_aba_lancamentos(wb_out, lancamentos, titulo, nome_aba):
    ws = wb_out.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value=titulo)
    apf(c, COR_TITULO_FUNDO); apft(c, bold=True, size=12, cor=COR_TITULO_FONTE)
    aln(c, h="center"); ws.row_dimensions[1].height = 26

    hdr(ws, 2, ["DATA", "CATEGORIA", "GRUPO DRE",
                "SITUACAO", "VALOR (R$)", "TIPO", "RESUMO"])
    ws.row_dimensions[3].height = 5

    ordenados = sorted(lancamentos,
                       key=lambda x: (x["data"] or date.min, x["categoria"]))
    for i, l in enumerate(ordenados, start=4):
        fundo = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR
        tipo  = "CREDITO" if l["valor"] >= 0 else "DEBITO"

        def fc(col, val, h="left", wrap=False):
            c = ws.cell(row=i, column=col, value=val)
            apf(c, fundo); apft(c, size=9); aln(c, h=h, wrap=wrap)
            c.border = BORDA_FINA
            return c

        fc(1, l["data_str"], h="center"); fc(2, l["categoria"])
        fc(3, l["rotulo_dre"]); fc(4, l["situacao"], h="center")
        cv = ws.cell(row=i, column=5, value=l["valor"])
        cv.number_format = FMT_MOEDA; apf(cv, fundo)
        cv.font = Font(bold=False, size=9, name="Calibri",
                       color=COR_POSITIVO if l["valor"] >= 0 else COR_NEGATIVO)
        cv.alignment = Alignment(horizontal="right", vertical="center")
        cv.border = BORDA_FINA
        ct = ws.cell(row=i, column=6, value=tipo); apf(ct, fundo)
        apft(ct, bold=True, size=9,
             cor="1A6B34" if tipo == "CREDITO" else "8B1A1A")
        aln(ct, h="center"); ct.border = BORDA_FINA
        fc(7, l["resumo"].replace("\n", " | "), wrap=True)
        ws.row_dimensions[i].height = 15

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 60


# ─────────────────────────────────────────────────────────────────────
# 10. PROCESSAMENTO DE UM MÊS
# ─────────────────────────────────────────────────────────────────────

def processar_mes(mes: int, ano: int, caminho: Path, hist: dict):
    """Processa um único arquivo mensal e atualiza historico.json."""

    legado = (mes == -1)  # modo compatível v1

    # Período legível
    if legado:
        # detecta mês/ano do conteúdo
        wb_tmp = openpyxl.load_workbook(str(caminho))
        ws_tmp = wb_tmp[ABA_ENTRADA] if ABA_ENTRADA in wb_tmp.sheetnames \
            else wb_tmp.active
        datas = []
        for row in ws_tmp.iter_rows(min_row=3, values_only=True):
            d = parse_data(row[0])
            if d:
                datas.append(d)
        if datas:
            mes = min(datas).month
            ano = min(datas).year
        else:
            mes, ano = datetime.now().month, datetime.now().year

    periodo_label = f"{MESES_PT[mes]} / {ano}"
    chave         = f"{ano}-{mes:02d}"

    print(f"\n  Processando: {caminho.name}  →  {periodo_label}")

    # Backup (apenas para arquivos não-legado ou se quiser)
    if not legado:
        bk = fazer_backup(caminho, mes, ano)
        print(f"    Backup: {bk.name}")

    # Saldo inicial do mês anterior
    saldo_inicial = obter_saldo_anterior(hist, mes, ano)
    if saldo_inicial:
        print(f"    Saldo transportado do mês anterior: R$ {saldo_inicial:,.2f}")
    else:
        print(f"    Saldo inicial: R$ 0,00 (sem histórico anterior)")

    # Saldo final da planilha
    saldo_final = obter_saldo_final_planilha(caminho, ABA_ENTRADA)

    # Lê e classifica lançamentos
    efetivados, em_aberto, nao_mapeados = ler_lancamentos(str(caminho), ABA_ENTRADA)
    print(f"    {len(efetivados)} efetivados | {len(em_aberto)} em aberto")

    if nao_mapeados:
        print(f"    ATENCAO — {len(nao_mapeados)} categoria(s) nao mapeada(s):")
        for c in sorted(nao_mapeados):
            print(f"      * {c}")

    # Totais e subtotais
    totais = defaultdict(float)
    for l in efetivados:
        if l["grupo_dre"] != "NAO_MAPEADO":
            totais[l["grupo_dre"]] += l["valor"]

    sub = calcular_subtotais(dict(totais))
    resultado_acum = calcular_acumulado_ano(hist, mes, ano, sub["resultado_final"])
    tania_stats = extrair_resumo_tania(efetivados, em_aberto)

    # Pasta de saída
    dir_ano = DIR_SAIDA / str(ano)
    dir_ano.mkdir(parents=True, exist_ok=True)
    arq_saida = dir_ano / f"DRE_Balancete_{mes:02d}{ano}.xlsx"

    # Gera planilha
    wb_out = openpyxl.Workbook()
    if "Sheet" in wb_out.sheetnames:
        del wb_out["Sheet"]

    gerar_aba_dre(wb_out, efetivados, periodo_label, nao_mapeados,
                  saldo_inicial=saldo_inicial, saldo_final=saldo_final,
                  resultado_acumulado=resultado_acum, mes=mes, ano=ano,
                  tania_stats=tania_stats)
    gerar_aba_balancete(wb_out, efetivados, periodo_label)
    gerar_aba_lancamentos(wb_out, efetivados,
                          f"LANCAMENTOS EFETIVADOS — {periodo_label}",
                          "Lancamentos Efetivados")
    gerar_aba_lancamentos(wb_out, em_aberto,
                          f"LANCAMENTOS EM ABERTO — {periodo_label}",
                          "Em Aberto")
    wb_out.save(str(arq_saida))
    print(f"    Salvo: {arq_saida}")

    # Atualiza historico.json
    hist.setdefault("meses", {})[chave] = {
        "periodo_label"       : periodo_label,
        "mes"                 : mes,
        "ano"                 : ano,
        "saldo_inicial"       : round(saldo_inicial, 2),
        "saldo_final"         : round(saldo_final, 2),
        "totais"              : {k: round(v, 2) for k, v in totais.items()},
        "subtotais"           : sub,
        "resultado_acumulado_ano": round(resultado_acum, 2),
        "qtd_efetivados"      : len(efetivados),
        "qtd_em_aberto"       : len(em_aberto),
        "em_aberto_detalhes"  : [
            {
                "data"      : l["data_str"],
                "categoria" : l["categoria"],
                "valor"     : l["valor"],
                "resumo"    : l["resumo"][:100],
            }
            for l in em_aberto
        ],
        "nao_mapeados"        : sorted(nao_mapeados),
        "arquivo_origem"      : caminho.name,
        "arq_saida"           : str(arq_saida),
        "gerado_em"           : datetime.now().isoformat(),
    }
    salvar_historico(hist)

    print(f"    Resultado do mes: R$ {sub['resultado_final']:,.2f}")
    print(f"    Saldo bancario final: R$ {saldo_final:,.2f}")
    print(f"    Acumulado no ano: R$ {resultado_acum:,.2f}")

    return sub["resultado_final"], saldo_final


# ─────────────────────────────────────────────────────────────────────
# 11. MAIN
# ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  EXTRATOR DRE / BALANCETE  v2.0")
    print("=" * 60)

    criar_estrutura()
    hist     = ler_historico()
    arquivos = descobrir_arquivos(hist)

    if not arquivos:
        print("\nNenhum arquivo novo para processar.")
        print(f"Coloque seus arquivos em:  {DIR_ENTRADA.resolve()}")
        print("Nomeie-os como:  controlefinanceiro_MMAAAA.xlsx")
        print("Exemplo:         controlefinanceiro_082026.xlsx")
        sys.exit(0)

    print(f"\n{len(arquivos)} arquivo(s) para processar:")
    for mes, ano, p in arquivos:
        lbl = f"{MESES_PT.get(mes, '?')} / {ano}" if mes > 0 else p.name
        print(f"  • {p.name}  ({lbl})")

    processados = 0
    for mes, ano, caminho in arquivos:
        processar_mes(mes, ano, caminho, hist)
        hist = ler_historico()   # recarrega após cada atualização
        processados += 1

    print(f"\n{'=' * 60}")
    print(f"  {processados} mes(es) processado(s) com sucesso.")
    print(f"  Historico salvo em: {HIST_PATH.resolve()}")
    print(f"  DREs disponíveis em: {DIR_SAIDA.resolve()}")
    print(f"\n  Para o consolidado anual execute:")
    print(f"  python3 consolidador_anual.py --ano {ano if ano > 0 else 2026}")
    print("=" * 60)


if __name__ == "__main__":
    main()
